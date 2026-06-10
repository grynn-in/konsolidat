"""Generic Consolidation Report — reads all config from ClickHouse.

Produces a multi-sheet Excel workbook:
  1. P&L Entity        — Before consolidation (entity × quarter)
  2. P&L Consolidated  — Waterfall: Entity → IC → CTA → Topside → Consolidated
  3. BS Entity         — Before consolidation (cumulative balances)
  4. BS Consolidated   — Waterfall
  5. Bridge            — High-level summary waterfall
  6. Diagnostics       — Litmus tests with pass/warn/fail

Optional (--include-epm):   EPM formula versions of sheets 1-5
Optional (--include-delta): Cross-sheet delta checks

Usage:
  python3 scripts/build_consolidation_report.py \
    --group GROUP_CORP --year 2024 \
    --output excel/Consolidation_Report.xlsx \
    [--include-epm] [--include-delta]
"""

import argparse
import json
import os
import sys
import urllib.request
import urllib.parse
from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── ClickHouse connection ────────────────────────────────────

def ch_query(sql, cfg=None):
    """Execute a ClickHouse HTTP query; return list of dicts."""
    host = os.environ.get("CLICKHOUSE_HOST", "localhost")
    port = os.environ.get("CLICKHOUSE_HTTP_PORT", "8123")
    user = os.environ.get("CLICKHOUSE_USER", "default")
    password = os.environ.get("CLICKHOUSE_PASSWORD", "")
    params = urllib.parse.urlencode({"user": user, "password": password})
    url = f"http://{host}:{port}/?{params}"
    req = urllib.request.Request(url, data=sql.encode("utf-8"))
    resp = urllib.request.urlopen(req)
    raw = resp.read().decode("utf-8")
    data = json.loads(raw)
    return data.get("data", [])


# ── Configuration dataclass ──────────────────────────────────

@dataclass
class ConsolidationConfig:
    group: str
    year: int
    reporting_currency: str = "USD"
    entities: list = field(default_factory=list)        # [{"data_area_id", "entity_name", "ownership_pct", "accounting_currency", "consolidation_method"}]
    pnl_sections: list = field(default_factory=list)    # [(account_type, [(main_account, account_name), ...])]
    bs_sections: list = field(default_factory=list)
    quarters: list = field(default_factory=lambda: ["Q1", "Q2", "Q3", "Q4"])
    q_end_period: dict = field(default_factory=lambda: {"Q1": 3, "Q2": 6, "Q3": 9, "Q4": 12})


def discover_config(group, year):
    """Build ConsolidationConfig from ClickHouse metadata."""
    # Get entities in this consolidation group
    rows = ch_query(f"""
        SELECT cg.consolidation_group, cg.data_area_id, cg.entity_name,
               cg.ownership_pct, cg.reporting_currency, cg.consolidation_method,
               le.accounting_currency
        FROM epm_gold.consolidation_groups AS cg
        LEFT JOIN epm_gold.silver_legal_entities AS le
            ON cg.data_area_id = le.data_area
        WHERE cg.consolidation_group = '{group}'
          AND cg.data_area_id != ''
          AND cg.consolidation_method = 'full'
        ORDER BY cg.data_area_id
        FORMAT JSON
    """)
    if not rows:
        print(f"ERROR: No entities found for group '{group}'")
        sys.exit(1)

    entities = []
    reporting_currency = "USD"
    for r in rows:
        reporting_currency = r.get("reporting_currency", "USD")
        entities.append({
            "data_area_id": r["data_area_id"],
            "entity_name": r["entity_name"],
            "ownership_pct": float(r["ownership_pct"]) / 100.0,
            "accounting_currency": r.get("accounting_currency") or reporting_currency,
            "consolidation_method": r["consolidation_method"],
        })

    entity_ids = [e["data_area_id"] for e in entities]
    entity_list_sql = ", ".join(f"'{e}'" for e in entity_ids)

    # Auto-detect year if not provided
    if year is None:
        yr_rows = ch_query(f"""
            SELECT max(fiscal_year) as max_year
            FROM epm_gold.gold_trial_balance
            WHERE data_area_id IN ({entity_list_sql})
              AND period_net_amount != 0
            FORMAT JSON
        """)
        year = int(yr_rows[0]["max_year"]) if yr_rows else 2024
        print(f"  Auto-detected fiscal year: {year}")

    # Discover P&L sections from data
    pnl_sections = _discover_sections(entity_list_sql, year, is_pnl=True)
    bs_sections = _discover_sections(entity_list_sql, year, is_pnl=False)

    return ConsolidationConfig(
        group=group,
        year=year,
        reporting_currency=reporting_currency,
        entities=entities,
        pnl_sections=pnl_sections,
        bs_sections=bs_sections,
    )


# ── P&L sub-section classification ──────────────────────────

# Map account_type_name → fine-grained sub-sections based on account ranges
PNL_SUBSECTIONS = [
    ("Revenue", "Revenue", lambda a: True),
    ("Expense", "Cost of Goods Sold", lambda a: a[:1] == "5" or a.startswith("6112")),
    ("Expense", "Operating Expenses", lambda a: a[:1] == "6" and not a.startswith("6112")),
    ("Profit and loss", "Other Income / Expense", lambda a: a[:1] in ("7", "8") and not a.startswith("802")),
    ("Profit and loss", "Income Tax", lambda a: a.startswith("802")),
    ("Expense", "Other Income / Expense", lambda a: a[:1] in ("7", "8")),
]

BS_SUBSECTIONS = [
    ("Asset", "Current Assets", lambda a: a[:2] in ("11", "12", "13", "14") and a[:3] not in ("120", "134")),
    ("Asset", "Non-Current Assets", lambda a: a[:2] in ("15", "16", "17", "18") or a[:3] in ("120", "134")),
    ("Balance sheet", "Current Assets", lambda a: a[:1] == "1"),
    ("Balance sheet", "Non-Current Assets", lambda a: a[:1] != "1"),
    ("Liability", "Current Liabilities", lambda a: a[:2] in ("20", "21", "22", "23")),
    ("Liability", "Non-Current Liabilities", lambda a: a[:2] in ("24", "25", "26", "27", "28", "29")),
    ("Equity", "Shareholders' Equity", lambda a: True),
]


def _classify_pnl_subsection(account_type_name, main_account):
    """Classify a P&L account into a fine-grained sub-section."""
    for atype, subsection, predicate in PNL_SUBSECTIONS:
        if account_type_name == atype and predicate(main_account):
            return subsection
    # Fallback: use account_type_name directly
    return account_type_name


def _classify_bs_subsection(account_type_name, main_account):
    """Classify a BS account into a fine-grained sub-section."""
    for atype, subsection, predicate in BS_SUBSECTIONS:
        if account_type_name == atype and predicate(main_account):
            return subsection
    return account_type_name


def _discover_sections(entity_list_sql, year, is_pnl=True):
    """Discover account sections dynamically from gold_trial_balance."""
    flag = "is_pnl = 1" if is_pnl else "is_balance_sheet = 1"
    rows = ch_query(f"""
        SELECT DISTINCT main_account, account_name, account_type_name
        FROM epm_gold.gold_trial_balance
        WHERE fiscal_year = {year}
          AND data_area_id IN ({entity_list_sql})
          AND {flag}
          AND period_net_amount != 0
        ORDER BY main_account
        FORMAT JSON
    """)

    # Group into sub-sections
    classify = _classify_pnl_subsection if is_pnl else _classify_bs_subsection
    subsection_map = defaultdict(list)  # subsection_name → [(main_account, account_name)]
    for r in rows:
        sub = classify(r["account_type_name"], r["main_account"])
        subsection_map[sub].append((r["main_account"], r["account_name"]))

    # Sort accounts within each sub-section
    for sub in subsection_map:
        subsection_map[sub].sort(key=lambda x: x[0])

    # Define ordering
    if is_pnl:
        section_order = ["Revenue", "Cost of Goods Sold", "Operating Expenses",
                         "Other Income / Expense", "Income Tax"]
        formulas = [
            (2, "Gross Profit", "SUM_AFTER", ["Revenue", "Cost of Goods Sold"]),
            (4, "Operating Income", "DIFF", ["Gross Profit", "Operating Expenses"]),
            (5, "Net Income Before Tax", "SUM", ["Operating Income", "Other Income / Expense"]),
            (6, "Net Income", "DIFF", ["Net Income Before Tax", "Income Tax"]),
        ]
    else:
        section_order = ["Current Assets", "Non-Current Assets",
                         "Current Liabilities", "Non-Current Liabilities",
                         "Shareholders' Equity"]
        formulas = [
            (2, "Total Assets", "SUM", ["Current Assets", "Non-Current Assets"]),
            (4, "Total Liabilities", "SUM", ["Current Liabilities", "Non-Current Liabilities"]),
            (5, "Total Liabilities & Equity", "SUM", ["Total Liabilities", "Shareholders' Equity"]),
            (5, "BS Check (A - L&E)", "DIFF", ["Total Assets", "Total Liabilities & Equity"]),
        ]

    # Add any discovered sub-sections not in our ordered list
    for sub in subsection_map:
        if sub not in section_order:
            section_order.append(sub)

    return {
        "sections": [(s, subsection_map.get(s, [])) for s in section_order if subsection_map.get(s)],
        "formulas": formulas,
    }


# ── Data fetching ────────────────────────────────────────────

def fetch_entity_pnl(cfg):
    """Fetch P&L data: entity × account × quarter from gold_trial_balance."""
    entity_ids = ", ".join(f"'{e['data_area_id']}'" for e in cfg.entities)
    rows = ch_query(f"""
        SELECT
            data_area_id,
            main_account,
            multiIf(
                fiscal_period IN (1,2,3), 'Q1',
                fiscal_period IN (4,5,6), 'Q2',
                fiscal_period IN (7,8,9), 'Q3',
                'Q4'
            ) as quarter,
            sum(period_net_amount) as amount
        FROM epm_gold.gold_trial_balance
        WHERE fiscal_year = {cfg.year}
          AND data_area_id IN ({entity_ids})
          AND is_pnl = 1
        GROUP BY data_area_id, main_account, quarter
        FORMAT JSON
    """)
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        result[r["data_area_id"]][r["main_account"]][r["quarter"]] += float(r["amount"])
    return result


def fetch_entity_bs(cfg):
    """Fetch BS data: entity × account × quarter-end from gold_balance_sheet."""
    entity_ids = ", ".join(f"'{e['data_area_id']}'" for e in cfg.entities)
    rows = ch_query(f"""
        SELECT
            data_area_id,
            main_account,
            fiscal_period,
            sum(cumulative_balance) as amount
        FROM epm_gold.gold_balance_sheet
        WHERE fiscal_year = {cfg.year}
          AND fiscal_period IN (3, 6, 9, 12)
          AND data_area_id IN ({entity_ids})
        GROUP BY data_area_id, main_account, fiscal_period
        FORMAT JSON
    """)
    period_to_q = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        q = period_to_q[int(r["fiscal_period"])]
        result[r["data_area_id"]][r["main_account"]][q] += float(r["amount"])
    return result


def fetch_consolidated_data(cfg):
    """Fetch waterfall data from gold_fully_consolidated_tb, grouped by adjustment_type."""
    rows = ch_query(f"""
        SELECT
            adjustment_type,
            main_account,
            multiIf(
                fiscal_period IN (1,2,3), 'Q1',
                fiscal_period IN (4,5,6), 'Q2',
                fiscal_period IN (7,8,9), 'Q3',
                'Q4'
            ) as quarter,
            sum(amount) as total_amount
        FROM epm_gold.gold_fully_consolidated_tb
        WHERE consolidation_group = '{cfg.group}'
          AND fiscal_year = {cfg.year}
        GROUP BY adjustment_type, main_account, quarter
        FORMAT JSON
    """)
    # {adjustment_type: {account: {quarter: amount}}}
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        result[r["adjustment_type"]][r["main_account"]][r["quarter"]] += float(r["total_amount"])
    return result


def fetch_consolidated_bs(cfg):
    """Fetch consolidated BS waterfall — cumulative amounts need special handling.

    For BS, we need cumulative (YTD) balances at quarter-end, not period sums.
    gold_fully_consolidated_tb has period amounts, so we cumulate them.
    """
    rows = ch_query(f"""
        SELECT
            adjustment_type,
            main_account,
            fiscal_period,
            sum(amount) as total_amount
        FROM epm_gold.gold_fully_consolidated_tb
        WHERE consolidation_group = '{cfg.group}'
          AND fiscal_year = {cfg.year}
        GROUP BY adjustment_type, main_account, fiscal_period
        ORDER BY adjustment_type, main_account, fiscal_period
        FORMAT JSON
    """)
    # Build period-level data first
    period_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in rows:
        period_data[r["adjustment_type"]][r["main_account"]][int(r["fiscal_period"])] += float(r["total_amount"])

    # Cumulate to quarter-end
    result = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for adj_type in period_data:
        for acct in period_data[adj_type]:
            cumul = 0.0
            for p in range(1, 13):
                cumul += period_data[adj_type][acct].get(p, 0.0)
                if p == 3:
                    result[adj_type][acct]["Q1"] = cumul
                elif p == 6:
                    result[adj_type][acct]["Q2"] = cumul
                elif p == 9:
                    result[adj_type][acct]["Q3"] = cumul
                elif p == 12:
                    result[adj_type][acct]["Q4"] = cumul
    return result


def fetch_nci_data(cfg):
    """Fetch NCI amounts from gold_consolidated_trial_balance."""
    rows = ch_query(f"""
        SELECT
            main_account,
            is_pnl,
            is_balance_sheet,
            multiIf(
                fiscal_period IN (1,2,3), 'Q1',
                fiscal_period IN (4,5,6), 'Q2',
                fiscal_period IN (7,8,9), 'Q3',
                'Q4'
            ) as quarter,
            sum(nci_amount) as nci_total
        FROM epm_gold.gold_consolidated_trial_balance
        WHERE consolidation_group = '{cfg.group}'
          AND fiscal_year = {cfg.year}
        GROUP BY main_account, is_pnl, is_balance_sheet, quarter
        FORMAT JSON
    """)
    pnl_nci = defaultdict(lambda: defaultdict(float))
    bs_nci = defaultdict(lambda: defaultdict(float))
    for r in rows:
        acct = r["main_account"]
        q = r["quarter"]
        nci = float(r["nci_total"])
        if int(r["is_pnl"]):
            pnl_nci[acct][q] += nci
        if int(r["is_balance_sheet"]):
            bs_nci[acct][q] += nci
    return pnl_nci, bs_nci


def fetch_nci_bs_cumulative(cfg):
    """Fetch NCI for BS — needs cumulative amounts at quarter-end."""
    rows = ch_query(f"""
        SELECT
            main_account,
            fiscal_period,
            sum(nci_amount) as nci_total
        FROM epm_gold.gold_consolidated_trial_balance
        WHERE consolidation_group = '{cfg.group}'
          AND fiscal_year = {cfg.year}
          AND is_balance_sheet = 1
        GROUP BY main_account, fiscal_period
        ORDER BY main_account, fiscal_period
        FORMAT JSON
    """)
    period_data = defaultdict(lambda: defaultdict(float))
    for r in rows:
        period_data[r["main_account"]][int(r["fiscal_period"])] += float(r["nci_total"])

    result = defaultdict(lambda: defaultdict(float))
    for acct in period_data:
        cumul = 0.0
        for p in range(1, 13):
            cumul += period_data[acct].get(p, 0.0)
            if p == 3:
                result[acct]["Q1"] = cumul
            elif p == 6:
                result[acct]["Q2"] = cumul
            elif p == 9:
                result[acct]["Q3"] = cumul
            elif p == 12:
                result[acct]["Q4"] = cumul
    return result


def compute_net_income_plug(bs_data, entities):
    """NI = -(sum of all BS accounts) per entity per quarter — the balancing plug."""
    result = defaultdict(dict)
    for ent_cfg in entities:
        eid = ent_cfg["data_area_id"]
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            total = sum(
                bs_data.get(eid, {}).get(acct, {}).get(q, 0)
                for acct in bs_data.get(eid, {})
            )
            result[eid][q] = -total
    return result


# ── Styles ───────────────────────────────────────────────────

DARK_BLUE = "1F3864"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
DARK_GREEN = "1D4E2E"
MED_GREEN = "2E7D46"
LIGHT_GREEN = "D6EFD6"
DARK_PURPLE = "3B1F6E"
MED_PURPLE = "5C3D99"
LIGHT_PURPLE = "E0D6F0"
DARK_TEAL = "1B4D5C"
MED_TEAL = "2A7A8C"
LIGHT_TEAL = "D0ECF0"
DARK_RED = "8B0000"
MED_RED = "C04040"
LIGHT_RED = "F2DCDB"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "C00000"

font_title = Font(name="Aptos", size=14, bold=True, color=WHITE)
font_header = Font(name="Aptos", size=10, bold=True, color=WHITE)
font_section = Font(name="Aptos", size=10, bold=True)
font_subtotal = Font(name="Aptos", size=10, bold=True)
font_total = Font(name="Aptos", size=11, bold=True)
font_normal = Font(name="Aptos", size=10)
font_param_label = Font(name="Aptos", size=10, bold=True)
font_param_value = Font(name="Aptos", size=10, bold=True)

thin_border = Border(bottom=Side(style="thin", color="B0B0B0"))
subtotal_border = Border(top=Side(style="thin", color=BLACK), bottom=Side(style="thin", color=BLACK))
total_border = Border(top=Side(style="thin", color=BLACK), bottom=Side(style="double", color=BLACK))

NUM_FMT = '#,##0'


def _make_theme(primary_dark, primary_med, primary_light):
    return {
        "fill_title": PatternFill(start_color=primary_dark, end_color=primary_dark, fill_type="solid"),
        "fill_header": PatternFill(start_color=primary_med, end_color=primary_med, fill_type="solid"),
        "fill_subtotal": PatternFill(start_color=primary_light, end_color=primary_light, fill_type="solid"),
        "fill_total": PatternFill(start_color=primary_light, end_color=primary_light, fill_type="solid"),
        "fill_alt": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
        "font_section": Font(name="Aptos", size=10, bold=True, color=primary_dark),
        "font_total": Font(name="Aptos", size=11, bold=True, color=primary_dark),
        "font_param_label": Font(name="Aptos", size=10, bold=True, color=primary_dark),
    }


THEME_PNL = _make_theme(DARK_BLUE, MED_BLUE, LIGHT_BLUE)
THEME_BS = _make_theme(DARK_GREEN, MED_GREEN, LIGHT_GREEN)
THEME_CONSOL = _make_theme(DARK_PURPLE, MED_PURPLE, LIGHT_PURPLE)
THEME_BRIDGE = _make_theme(DARK_TEAL, MED_TEAL, LIGHT_TEAL)
THEME_DELTA = _make_theme(DARK_RED, MED_RED, LIGHT_RED)
THEME_DIAG = _make_theme("2D572C", "3A7D44", "D6EFD6")  # green for diagnostics


# ── Sheet builder helpers ────────────────────────────────────

def _write_title(ws, text, theme, total_cols):
    ws.merge_cells("A1:B1")
    ws["A1"] = text
    ws["A1"].font = font_title
    ws["A1"].fill = theme["fill_title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).fill = theme["fill_title"]


def _write_params(ws, params, theme):
    """Write parameter rows (A3:B3, A4:B4, etc.)."""
    for i, (label, value) in enumerate(params):
        r = 3 + i
        ws.cell(row=r, column=1, value=label).font = theme["font_param_label"]
        ws.cell(row=r, column=2, value=value).font = font_param_value


def _get_all_accounts(section_def):
    """Extract flat list of (main_account, account_name) from section definition."""
    accounts = []
    for _section_name, accts in section_def["sections"]:
        accounts.extend(accts)
    return accounts


# ── Entity sheet builder ─────────────────────────────────────

def build_entity_sheet(ws, cfg, section_def, entity_data, theme,
                       sign_flip=True, net_income=None, is_bs=False):
    """Build entity-level sheet with columns for each entity × quarter."""
    t = theme
    entities = cfg.entities
    quarters = cfg.quarters
    total_cols = 2 + len(entities) * len(quarters)

    _write_title(ws, "P&L — Before Consolidation" if not is_bs else "Balance Sheet — Before Consolidation", t, total_cols)
    _write_params(ws, [
        ("Fiscal Year", f"FY{cfg.year}"),
        ("Group", f"{cfg.group} ({cfg.reporting_currency})"),
    ], t)

    # Column headers — two rows: entity name + quarter
    header_row1 = 6
    header_row2 = 7
    for c_idx in (1, 2):
        ws.cell(row=header_row1, column=c_idx).fill = t["fill_header"]
        ws.cell(row=header_row2, column=c_idx).fill = t["fill_header"]
    ws.cell(row=header_row1, column=1, value="Account").font = font_header
    ws.cell(row=header_row1, column=2, value="Description").font = font_header

    col = 3
    for ent in entities:
        ccy = ent["accounting_currency"]
        start_col = col
        for q in quarters:
            c = ws.cell(row=header_row2, column=col, value=q)
            c.font = font_header
            c.fill = t["fill_header"]
            c.alignment = Alignment(horizontal="center")
            col += 1
        ws.merge_cells(start_row=header_row1, start_column=start_col,
                       end_row=header_row1, end_column=col - 1)
        c = ws.cell(row=header_row1, column=start_col,
                    value=f"{ent['data_area_id']} ({ccy})")
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    sign = -1 if sign_flip else 1
    n_data_cols = len(entities) * len(quarters)

    # Track subtotal rows for formulas
    subtotal_rows = {}
    row = 8

    for section_name, accounts in section_def["sections"]:
        # Section header
        ws.cell(row=row, column=2, value=section_name).font = t["font_section"]
        row += 1
        section_data_rows = []

        for main_acct, acct_name in accounts:
            section_data_rows.append(row)
            ws.cell(row=row, column=1, value=main_acct).font = font_normal
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2, value=acct_name).font = font_normal
            ws.cell(row=row, column=2).alignment = Alignment(indent=2)

            is_odd = len(section_data_rows) % 2 == 1
            if is_odd:
                ws.cell(row=row, column=1).fill = t["fill_alt"]
                ws.cell(row=row, column=2).fill = t["fill_alt"]

            col_idx = 3
            for ent in entities:
                eid = ent["data_area_id"]
                for q in quarters:
                    c = ws.cell(row=row, column=col_idx)
                    raw = entity_data.get(eid, {}).get(main_acct, {}).get(q, 0)
                    c.value = round(raw * sign) if raw else 0
                    c.number_format = NUM_FMT
                    c.font = font_normal
                    c.border = thin_border
                    if is_odd:
                        c.fill = t["fill_alt"]
                    col_idx += 1
            row += 1

        # NI plug row for BS
        if is_bs and section_name == "Shareholders' Equity" and net_income:
            section_data_rows.append(row)
            ws.cell(row=row, column=1, value="NI").font = font_normal
            ws.cell(row=row, column=2, value="Net Income (plug)").font = font_normal
            ws.cell(row=row, column=2).alignment = Alignment(indent=2)
            is_odd = len(section_data_rows) % 2 == 1
            if is_odd:
                ws.cell(row=row, column=1).fill = t["fill_alt"]
                ws.cell(row=row, column=2).fill = t["fill_alt"]
            col_idx = 3
            for ent in entities:
                eid = ent["data_area_id"]
                for q in quarters:
                    c = ws.cell(row=row, column=col_idx)
                    ni = net_income.get(eid, {}).get(q, 0)
                    c.value = round(ni * sign)
                    c.number_format = NUM_FMT
                    c.font = font_normal
                    c.border = thin_border
                    if is_odd:
                        c.fill = t["fill_alt"]
                    col_idx += 1
            row += 1

        # Subtotal row
        sub_label = f"Total {section_name}"
        ws.cell(row=row, column=2, value=sub_label).font = font_subtotal
        ws.cell(row=row, column=2).fill = t["fill_subtotal"]
        ws.cell(row=row, column=1).fill = t["fill_subtotal"]
        for ci in range(n_data_cols):
            col_idx = ci + 3
            c = ws.cell(row=row, column=col_idx)
            if section_data_rows:
                parts = [f"{get_column_letter(col_idx)}{r}" for r in section_data_rows]
                c.value = "=" + "+".join(parts)
            else:
                c.value = 0
            c.font = font_subtotal
            c.fill = t["fill_subtotal"]
            c.number_format = NUM_FMT
            c.border = subtotal_border
        subtotal_rows[sub_label] = row
        subtotal_rows[section_name] = row  # alias
        row += 1
        row += 1  # blank row

    # Formula rows
    for _, formula_label, formula_type, operands in section_def["formulas"]:
        ws.cell(row=row, column=2, value=formula_label).font = t["font_total"]
        ws.cell(row=row, column=2).fill = t["fill_total"]
        ws.cell(row=row, column=1).fill = t["fill_total"]
        is_check = "Check" in formula_label
        for ci in range(n_data_cols):
            col_idx = ci + 3
            cl = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx)
            c.value = _build_formula(formula_type, operands, subtotal_rows, cl)
            c.font = Font(name="Aptos", size=11, bold=True, color=RED) if is_check else t["font_total"]
            c.fill = t["fill_total"]
            c.number_format = NUM_FMT
            c.border = total_border if not is_check else subtotal_border
        subtotal_rows[formula_label] = row
        row += 1
        row += 1  # blank

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for ci in range(n_data_cols):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14
    ws.freeze_panes = "C8"


def _build_formula(formula_type, operands, subtotal_rows, col_letter):
    """Build an Excel formula from operand references."""
    refs = []
    for op in operands:
        # Try exact match first, then "Total <op>"
        r = subtotal_rows.get(op) or subtotal_rows.get(f"Total {op}")
        if r:
            refs.append(f"{col_letter}{r}")

    if not refs:
        return 0

    if formula_type == "SUM" or formula_type == "SUM_AFTER":
        return "=" + "+".join(refs)
    elif formula_type == "DIFF":
        if len(refs) >= 2:
            return f"={refs[0]}-{refs[1]}"
        return f"={refs[0]}"
    return 0


# ── Consolidated sheet builder ───────────────────────────────

WATERFALL_BLOCKS = [
    ("Entity", "entity"),
    ("IC Elim", "ic_elimination"),
    ("CTA", "cta"),
    ("Topside", "topside"),
    ("Consolidated", None),  # computed
]


def build_consolidated_sheet(ws, cfg, section_def, consol_data, nci_data,
                             theme, is_bs=False, sign_flip=True):
    """Build consolidated waterfall sheet.

    Columns: Account | Description | [block × Q1..Q4]...
    """
    t = theme
    quarters = cfg.quarters
    # Determine which blocks have data
    active_blocks = []
    for label, adj_type in WATERFALL_BLOCKS:
        if adj_type is None or adj_type in consol_data:
            active_blocks.append((label, adj_type))
    # Always include Entity, NCI if relevant, and Consolidated
    has_nci = any(
        nci_data.get(acct, {}).get(q, 0) != 0
        for acct in nci_data for q in quarters
    )
    # Build final block list: Entity, IC, CTA, Topside, NCI, Consolidated
    blocks = []
    for label, adj_type in active_blocks:
        if adj_type is not None:
            blocks.append((label, adj_type))
    if has_nci:
        blocks.append(("NCI", "_nci"))
    blocks.append(("Consolidated", None))

    n_blocks = len(blocks)
    total_data_cols = n_blocks * len(quarters)
    total_cols = 2 + total_data_cols

    title_text = "P&L — Consolidated" if not is_bs else "Balance Sheet — Consolidated"
    _write_title(ws, title_text, t, total_cols)
    _write_params(ws, [
        ("Fiscal Year", f"FY{cfg.year}"),
        ("Group", f"{cfg.group} ({cfg.reporting_currency})"),
    ], t)

    # Column headers
    header_row1 = 6
    header_row2 = 7
    for c_idx in (1, 2):
        ws.cell(row=header_row1, column=c_idx).fill = t["fill_header"]
        ws.cell(row=header_row2, column=c_idx).fill = t["fill_header"]
    ws.cell(row=header_row1, column=1, value="Account").font = font_header
    ws.cell(row=header_row1, column=2, value="Description").font = font_header

    col = 3
    for block_label, _ in blocks:
        start_col = col
        for q in quarters:
            c = ws.cell(row=header_row2, column=col, value=q)
            c.font = font_header
            c.fill = t["fill_header"]
            c.alignment = Alignment(horizontal="center")
            col += 1
        ws.merge_cells(start_row=header_row1, start_column=start_col,
                       end_row=header_row1, end_column=col - 1)
        c = ws.cell(row=header_row1, column=start_col, value=block_label)
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    n_per = len(quarters)
    block_starts = {}
    for bi, (bl, _) in enumerate(blocks):
        block_starts[bl] = 3 + bi * n_per

    sign = -1 if sign_flip else 1
    subtotal_rows = {}
    row = 8

    for section_name, accounts in section_def["sections"]:
        ws.cell(row=row, column=2, value=section_name).font = t["font_section"]
        row += 1
        section_data_rows = []

        for main_acct, acct_name in accounts:
            section_data_rows.append(row)
            ws.cell(row=row, column=1, value=main_acct).font = font_normal
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2, value=acct_name).font = font_normal
            ws.cell(row=row, column=2).alignment = Alignment(indent=2)

            is_odd = len(section_data_rows) % 2 == 1
            if is_odd:
                ws.cell(row=row, column=1).fill = t["fill_alt"]
                ws.cell(row=row, column=2).fill = t["fill_alt"]

            for qi, q in enumerate(quarters):
                for block_label, adj_type in blocks:
                    col_idx = block_starts[block_label] + qi
                    c = ws.cell(row=row, column=col_idx)

                    if adj_type is None:
                        # Consolidated = sum of all other blocks
                        parts = []
                        for bl2, _ in blocks:
                            if bl2 != "Consolidated":
                                parts.append(f"{get_column_letter(block_starts[bl2] + qi)}{row}")
                        c.value = "=" + "+".join(parts)
                    elif adj_type == "_nci":
                        val = nci_data.get(main_acct, {}).get(q, 0) * sign
                        c.value = round(val)
                    else:
                        val = consol_data.get(adj_type, {}).get(main_acct, {}).get(q, 0) * sign
                        c.value = round(val)

                    c.number_format = NUM_FMT
                    c.font = font_normal
                    c.border = thin_border
                    if is_odd:
                        c.fill = t["fill_alt"]
            row += 1

        # Subtotal
        sub_label = f"Total {section_name}"
        ws.cell(row=row, column=2, value=sub_label).font = font_subtotal
        ws.cell(row=row, column=2).fill = t["fill_subtotal"]
        ws.cell(row=row, column=1).fill = t["fill_subtotal"]
        for ci in range(total_data_cols):
            col_idx = ci + 3
            c = ws.cell(row=row, column=col_idx)
            if section_data_rows:
                parts = [f"{get_column_letter(col_idx)}{r}" for r in section_data_rows]
                c.value = "=" + "+".join(parts)
            else:
                c.value = 0
            c.font = font_subtotal
            c.fill = t["fill_subtotal"]
            c.number_format = NUM_FMT
            c.border = subtotal_border
        subtotal_rows[sub_label] = row
        subtotal_rows[section_name] = row
        row += 1
        row += 1  # blank

    # Formula rows
    for _, formula_label, formula_type, operands in section_def["formulas"]:
        ws.cell(row=row, column=2, value=formula_label).font = t["font_total"]
        ws.cell(row=row, column=2).fill = t["fill_total"]
        ws.cell(row=row, column=1).fill = t["fill_total"]
        is_check = "Check" in formula_label
        for ci in range(total_data_cols):
            col_idx = ci + 3
            cl = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx)
            c.value = _build_formula(formula_type, operands, subtotal_rows, cl)
            c.font = Font(name="Aptos", size=11, bold=True, color=RED) if is_check else t["font_total"]
            c.fill = t["fill_total"]
            c.number_format = NUM_FMT
            c.border = total_border if not is_check else subtotal_border
        subtotal_rows[formula_label] = row
        row += 1
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for ci in range(total_data_cols):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14
    ws.freeze_panes = "C8"


# ── Bridge sheet builder ─────────────────────────────────────

def build_bridge_sheet(ws, cfg, pnl_consol, bs_consol, pnl_nci, bs_nci):
    """Build high-level consolidation bridge summary."""
    t = THEME_BRIDGE
    quarters = cfg.quarters

    # Determine active adjustment types
    all_adj_types = set()
    for data in (pnl_consol, bs_consol):
        all_adj_types.update(data.keys())
    adj_types_ordered = []
    for adj in ["entity", "ic_elimination", "cta", "topside", "reclassification"]:
        if adj in all_adj_types:
            adj_types_ordered.append(adj)

    has_nci = any(
        pnl_nci.get(a, {}).get(q, 0) != 0 or bs_nci.get(a, {}).get(q, 0) != 0
        for a in set(list(pnl_nci.keys()) + list(bs_nci.keys()))
        for q in quarters
    )

    blocks = [(adj.replace("_", " ").title(), adj) for adj in adj_types_ordered]
    if has_nci:
        blocks.append(("NCI", "_nci"))
    blocks.append(("Consolidated", None))

    n_q = len(quarters)
    total_cols = 2 + len(blocks) * n_q

    _write_title(ws, "Consolidation Bridge", t, total_cols)
    _write_params(ws, [("Fiscal Year", f"FY{cfg.year}")], t)

    header_row1 = 5
    header_row2 = 6
    ws.cell(row=header_row2, column=1, value="Metric").font = font_header
    for c_idx in (1, 2):
        ws.cell(row=header_row1, column=c_idx).fill = t["fill_header"]
        ws.cell(row=header_row2, column=c_idx).fill = t["fill_header"]

    col = 3
    block_starts = {}
    for block_label, _ in blocks:
        block_starts[block_label] = col
        start_col = col
        for q in quarters:
            c = ws.cell(row=header_row2, column=col, value=q)
            c.font = font_header
            c.fill = t["fill_header"]
            c.alignment = Alignment(horizontal="center")
            col += 1
        ws.merge_cells(start_row=header_row1, start_column=start_col,
                       end_row=header_row1, end_column=col - 1)
        c = ws.cell(row=header_row1, column=start_col, value=block_label)
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    # Bridge metrics — derived from section subtotals
    def sum_adj(consol_data, nci_data, accounts, adj_type, quarter):
        total = 0
        if adj_type == "_nci":
            for a in accounts:
                total += nci_data.get(a, {}).get(quarter, 0)
        elif adj_type is not None:
            for a in accounts:
                total += consol_data.get(adj_type, {}).get(a, {}).get(quarter, 0)
        return total

    # Get all P&L and BS accounts from sections
    pnl_sections = cfg.pnl_sections
    bs_sections = cfg.bs_sections

    def accounts_for(section_def, section_names):
        accts = []
        for sname, saccts in section_def["sections"]:
            if sname in section_names:
                accts.extend([a for a, _ in saccts])
        return accts

    rev_accts = accounts_for(pnl_sections, ["Revenue"])
    cogs_accts = accounts_for(pnl_sections, ["Cost of Goods Sold"])
    opex_accts = accounts_for(pnl_sections, ["Operating Expenses"])
    asset_accts = accounts_for(bs_sections, ["Current Assets", "Non-Current Assets"])
    liab_accts = accounts_for(bs_sections, ["Current Liabilities", "Non-Current Liabilities"])
    equity_accts = accounts_for(bs_sections, ["Shareholders' Equity"])

    metrics = [
        ("Total Revenue", rev_accts, pnl_consol, pnl_nci, True),
        ("Total COGS", cogs_accts, pnl_consol, pnl_nci, True),
        ("Gross Profit", None, None, None, True),
        ("Total OpEx", opex_accts, pnl_consol, pnl_nci, True),
        ("Operating Income", None, None, None, True),
        ("_BLANK", None, None, None, None),
        ("Total Assets", asset_accts, bs_consol, bs_nci, True),
        ("Total Liabilities", liab_accts, bs_consol, bs_nci, True),
        ("Total Equity", equity_accts, bs_consol, bs_nci, True),
    ]

    row = 7
    metric_rows = {}
    for name, accts, consol_data_ref, nci_ref, do_sign_flip in metrics:
        if name == "_BLANK":
            row += 1
            continue

        ws.cell(row=row, column=1).fill = t["fill_subtotal"]
        ws.cell(row=row, column=2, value=name).font = font_subtotal
        ws.cell(row=row, column=2).fill = t["fill_subtotal"]
        metric_rows[name] = row

        is_computed = accts is None
        sign = -1 if do_sign_flip else 1

        for qi, q in enumerate(quarters):
            for block_label, adj_type in blocks:
                col_idx = block_starts[block_label] + qi
                c = ws.cell(row=row, column=col_idx)
                c.number_format = NUM_FMT
                c.font = font_subtotal
                c.border = thin_border

                if is_computed:
                    cl = get_column_letter(col_idx)
                    if name == "Gross Profit":
                        c.value = f"={cl}{metric_rows['Total Revenue']}-{cl}{metric_rows['Total COGS']}"
                    elif name == "Operating Income":
                        c.value = f"={cl}{metric_rows['Gross Profit']}-{cl}{metric_rows['Total OpEx']}"
                    else:
                        c.value = 0
                elif adj_type is None:
                    # Consolidated = sum of all other blocks
                    parts = []
                    for bl2, _ in blocks:
                        if bl2 != "Consolidated":
                            parts.append(f"{get_column_letter(block_starts[bl2] + qi)}{row}")
                    c.value = "=" + "+".join(parts)
                else:
                    val = sum_adj(consol_data_ref, nci_ref, accts, adj_type, q) * sign
                    c.value = round(val)
        row += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    for ci in range(len(blocks) * n_q):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14
    ws.freeze_panes = "C7"


# ── Diagnostics sheet builder ────────────────────────────────

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INFO_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _status_cell(ws, row, col, status, value=None):
    """Write a status cell with conditional formatting."""
    c = ws.cell(row=row, column=col)
    if status == "PASS":
        c.value = "PASS"
        c.fill = PASS_FILL
    elif status == "WARN":
        c.value = "WARN"
        c.fill = WARN_FILL
    elif status == "FAIL":
        c.value = "FAIL"
        c.fill = FAIL_FILL
    elif status == "INFO":
        c.value = value or "INFO"
        c.fill = INFO_FILL
    c.font = Font(name="Aptos", size=10, bold=True)
    c.alignment = Alignment(horizontal="center")


def build_diagnostics_sheet(ws, cfg, entity_pnl, entity_bs, consol_pnl, consol_bs):
    """Build Diagnostics sheet with litmus tests."""
    t = THEME_DIAG
    _write_title(ws, "Diagnostics", t, 6)

    # Headers
    row = 3
    headers = ["Category", "Test", "Expected", "Actual", "Status", "Diagnosis"]
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=ci + 1, value=h)
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    row = 4
    tests = []

    # ── Category 1: Data Coverage (informational) ────────────
    entity_ids = [e["data_area_id"] for e in cfg.entities]
    entity_list = ", ".join(entity_ids)
    tests.append(("Data Coverage", "Entities in group", "-", entity_list, "INFO", ""))

    # Entities with data
    entities_with_pnl = [e for e in entity_ids if entity_pnl.get(e)]
    entities_with_bs = [e for e in entity_ids if entity_bs.get(e)]
    all_with_data = sorted(set(entities_with_pnl + entities_with_bs))
    tests.append(("Data Coverage", "Entities with data", entity_list,
                  ", ".join(all_with_data) if all_with_data else "NONE",
                  "PASS" if set(entity_ids) == set(all_with_data) else "WARN",
                  "" if set(entity_ids) == set(all_with_data) else "Some entities have no GL data"))

    # P&L / BS account counts
    pnl_accts = set()
    bs_accts = set()
    for s_name, s_accts in cfg.pnl_sections["sections"]:
        pnl_accts.update(a for a, _ in s_accts)
    for s_name, s_accts in cfg.bs_sections["sections"]:
        bs_accts.update(a for a, _ in s_accts)
    tests.append(("Data Coverage", "P&L account count", "> 0", str(len(pnl_accts)),
                  "PASS" if pnl_accts else "FAIL", "" if pnl_accts else "No P&L accounts found"))
    tests.append(("Data Coverage", "BS account count", "> 0", str(len(bs_accts)),
                  "PASS" if bs_accts else "FAIL", "" if bs_accts else "No BS accounts found"))

    # ── Category 2: Ownership checks ─────────────────────────
    for ent in cfg.entities:
        pct = ent["ownership_pct"] * 100
        status = "PASS" if 0 < pct <= 100 else "FAIL"
        tests.append(("Ownership", f"{ent['data_area_id']} ownership",
                      "0 < pct <= 100", f"{pct:.0f}%", status,
                      "" if status == "PASS" else "Bad config in consolidation_groups"))

    # ── Category 3: BS Entity Check ──────────────────────────
    for ent in cfg.entities:
        eid = ent["data_area_id"]
        for q in cfg.quarters:
            bs_total = sum(
                entity_bs.get(eid, {}).get(acct, {}).get(q, 0)
                for acct in entity_bs.get(eid, {})
            )
            abs_total = abs(bs_total)
            if abs_total < 1000:
                status = "PASS"
            elif abs_total < 100000:
                status = "WARN"
            else:
                status = "FAIL"
            tests.append(("BS Entity Check", f"{eid} {q} BS residual",
                          "~0", f"{bs_total:,.0f}", status,
                          "" if status == "PASS" else "Missing accounts or unclosed income"))

    # ── Category 4: Consolidated layer additivity ────────────
    # Check entity layer in fully_consolidated == sum of group_amount from consolidated_trial_balance
    entity_layer_totals = {}
    for acct in consol_pnl.get("entity", {}):
        for q in cfg.quarters:
            key = (acct, q)
            entity_layer_totals[key] = entity_layer_totals.get(key, 0) + consol_pnl["entity"][acct].get(q, 0)

    # Simple check: does consolidated data have entity rows?
    entity_row_count = sum(1 for acct in consol_pnl.get("entity", {})
                          for q in cfg.quarters
                          if consol_pnl["entity"][acct].get(q, 0) != 0)
    tests.append(("Layer Additivity", "Entity layer has data",
                  "> 0", str(entity_row_count),
                  "PASS" if entity_row_count > 0 else "WARN",
                  "" if entity_row_count > 0 else "No entity rows in gold_fully_consolidated_tb"))

    # ── Category 5: IC Eliminations ──────────────────────────
    ic_count = 0
    for acct in consol_pnl.get("ic_elimination", {}):
        ic_count += sum(1 for q in cfg.quarters
                       if consol_pnl["ic_elimination"][acct].get(q, 0) != 0)
    tests.append(("IC Eliminations", "IC row count",
                  "> 0 (if IC rules exist)", str(ic_count),
                  "PASS" if ic_count > 0 else "WARN",
                  "" if ic_count > 0 else "No IC data — either no rules or no matching transactions"))

    # ── Category 6: CTA check ────────────────────────────────
    cta_count = sum(1 for acct in consol_pnl.get("cta", {})
                   for q in cfg.quarters
                   if consol_pnl["cta"][acct].get(q, 0) != 0)
    non_usd_entities = [e for e in cfg.entities if e["accounting_currency"] != cfg.reporting_currency]
    if non_usd_entities:
        tests.append(("FX / CTA", "CTA entries exist",
                      "> 0 (non-reporting-currency entities)", str(cta_count),
                      "PASS" if cta_count > 0 else "WARN",
                      "" if cta_count > 0 else "No CTA entries — check silver_exchange_rates"))
    else:
        tests.append(("FX / CTA", "CTA entries",
                      "0 (all entities in reporting currency)", str(cta_count),
                      "PASS", ""))

    # ── Category 7: FX Rates ─────────────────────────────────
    try:
        rate_rows = ch_query(f"""
            SELECT DISTINCT
                ctb.data_area_id,
                ctb.accounting_currency,
                ctb.reporting_currency,
                ctb.closing_rate,
                ctb.average_rate,
                ctb.translation_rate
            FROM epm_gold.gold_consolidated_trial_balance AS ctb
            WHERE ctb.consolidation_group = '{cfg.group}'
              AND ctb.fiscal_year = {cfg.year}
              AND ctb.accounting_currency != ctb.reporting_currency
            LIMIT 20
            FORMAT JSON
        """)
        for rr in rate_rows:
            rate = float(rr.get("translation_rate", 0))
            eid = rr["data_area_id"]
            from_ccy = rr["accounting_currency"]
            to_ccy = rr["reporting_currency"]
            status = "PASS" if rate != 1.0 and rate != 0 else "WARN"
            tests.append(("FX Rates", f"{eid} {from_ccy}→{to_ccy} rate",
                          "!= 1.0", f"{rate:.6f}", status,
                          "" if status == "PASS" else "Rate defaulted — check silver_exchange_rates"))
    except Exception:
        tests.append(("FX Rates", "Rate query", "success", "FAILED", "WARN", "Could not query rates"))

    # Write all tests
    for category, test_name, expected, actual, status, diagnosis in tests:
        ws.cell(row=row, column=1, value=category).font = font_normal
        ws.cell(row=row, column=2, value=test_name).font = font_normal
        ws.cell(row=row, column=3, value=expected).font = font_normal
        ws.cell(row=row, column=4, value=actual).font = font_normal
        _status_cell(ws, row, 5, status)
        ws.cell(row=row, column=6, value=diagnosis).font = Font(name="Aptos", size=9, italic=True, color="666666")
        row += 1

    # Summary row
    row += 1
    pass_count = sum(1 for t in tests if t[4] == "PASS")
    warn_count = sum(1 for t in tests if t[4] == "WARN")
    fail_count = sum(1 for t in tests if t[4] == "FAIL")
    info_count = sum(1 for t in tests if t[4] == "INFO")
    ws.cell(row=row, column=1, value="Summary").font = font_subtotal
    ws.cell(row=row, column=2,
            value=f"PASS: {pass_count}  WARN: {warn_count}  FAIL: {fail_count}  INFO: {info_count}").font = font_subtotal

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 50
    ws.freeze_panes = "A4"


# ── EPM formula sheet builders ───────────────────────────────

def build_entity_sheet_epm(ws, cfg, section_def, theme, measure, periods,
                           sign_flip=True, is_bs=False):
    """Build entity sheet using =EPM() formulas."""
    t = theme
    entities = cfg.entities
    sign = "-" if sign_flip else ""
    total_cols = 2 + len(entities) * len(periods)

    title = ("P&L — Entity (EPM)" if not is_bs else "BS — Entity (EPM)")
    _write_title(ws, title, t, total_cols)
    ws["A3"] = "Fiscal Year"
    ws["A3"].font = t["font_param_label"]
    ws["B3"] = cfg.year
    ws["B3"].font = font_param_value
    year_ref = "$B$3"

    ws["A4"] = "Measure"
    ws["A4"].font = t["font_param_label"]
    ws["B4"] = measure
    ws["B4"].font = font_param_value

    # Entity codes in row 5
    col = 3
    entity_refs = {}
    for ent in entities:
        cl = get_column_letter(col)
        ws.cell(row=5, column=col, value=ent["data_area_id"])
        ws.cell(row=5, column=col).font = Font(name="Aptos", size=8, color="999999")
        entity_refs[ent["data_area_id"]] = f"${cl}$5"
        col += len(periods)

    header_row1 = 6
    header_row2 = 7
    for c_idx in (1, 2):
        ws.cell(row=header_row1, column=c_idx).fill = t["fill_header"]
        ws.cell(row=header_row2, column=c_idx).fill = t["fill_header"]
    ws.cell(row=header_row1, column=1, value="Account").font = font_header
    ws.cell(row=header_row1, column=2, value="Description").font = font_header

    col = 3
    for ent in entities:
        ccy = ent["accounting_currency"]
        start_col = col
        for pval, plabel in periods:
            c = ws.cell(row=header_row2, column=col, value=plabel)
            c.font = font_header
            c.fill = t["fill_header"]
            c.alignment = Alignment(horizontal="center")
            col += 1
        ws.merge_cells(start_row=header_row1, start_column=start_col,
                       end_row=header_row1, end_column=col - 1)
        c = ws.cell(row=header_row1, column=start_col, value=f"{ent['data_area_id']} ({ccy})")
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    n_data_cols = len(entities) * len(periods)
    subtotal_rows = {}
    row = 8

    for section_name, accounts in section_def["sections"]:
        ws.cell(row=row, column=2, value=section_name).font = t["font_section"]
        row += 1
        section_data_rows = []

        for main_acct, acct_name in accounts:
            section_data_rows.append(row)
            ws.cell(row=row, column=1, value=main_acct).font = font_normal
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2, value=acct_name).font = font_normal
            ws.cell(row=row, column=2).alignment = Alignment(indent=2)

            is_odd = len(section_data_rows) % 2 == 1
            if is_odd:
                ws.cell(row=row, column=1).fill = t["fill_alt"]
                ws.cell(row=row, column=2).fill = t["fill_alt"]

            col_idx = 3
            for ent in entities:
                ent_ref = entity_refs[ent["data_area_id"]]
                for pval, plabel in periods:
                    c = ws.cell(row=row, column=col_idx)
                    parg = f'"{pval}"' if isinstance(pval, str) else str(pval)
                    c.value = f'={sign}EPM({ent_ref},{year_ref},{parg},$A{row},"{measure}")'
                    c.number_format = NUM_FMT
                    c.font = font_normal
                    c.border = thin_border
                    if is_odd:
                        c.fill = t["fill_alt"]
                    col_idx += 1
            row += 1

        sub_label = f"Total {section_name}"
        ws.cell(row=row, column=2, value=sub_label).font = font_subtotal
        ws.cell(row=row, column=2).fill = t["fill_subtotal"]
        ws.cell(row=row, column=1).fill = t["fill_subtotal"]
        for ci in range(n_data_cols):
            col_idx = ci + 3
            c = ws.cell(row=row, column=col_idx)
            if section_data_rows:
                parts = [f"{get_column_letter(col_idx)}{r}" for r in section_data_rows]
                c.value = "=" + "+".join(parts)
            else:
                c.value = 0
            c.font = font_subtotal
            c.fill = t["fill_subtotal"]
            c.number_format = NUM_FMT
            c.border = subtotal_border
        subtotal_rows[sub_label] = row
        subtotal_rows[section_name] = row
        row += 1
        row += 1

    for _, formula_label, formula_type, operands in section_def["formulas"]:
        ws.cell(row=row, column=2, value=formula_label).font = t["font_total"]
        ws.cell(row=row, column=2).fill = t["fill_total"]
        ws.cell(row=row, column=1).fill = t["fill_total"]
        is_check = "Check" in formula_label
        for ci in range(n_data_cols):
            col_idx = ci + 3
            cl = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx)
            c.value = _build_formula(formula_type, operands, subtotal_rows, cl)
            c.font = Font(name="Aptos", size=11, bold=True, color=RED) if is_check else t["font_total"]
            c.fill = t["fill_total"]
            c.number_format = NUM_FMT
            c.border = total_border if not is_check else subtotal_border
        subtotal_rows[formula_label] = row
        row += 1
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for ci in range(n_data_cols):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14
    ws.freeze_panes = "C8"


# ── Delta sheet builder ──────────────────────────────────────

def build_delta_sheet(ws, hardcoded_sheet, epm_sheet, section_def, num_data_cols):
    """Build delta sheet: Hardcoded - EPM for every data cell."""
    t = THEME_DELTA
    hc = f"'{hardcoded_sheet}'"
    ep = f"'{epm_sheet}'"

    _write_title(ws, f"Delta: {hardcoded_sheet}", t, num_data_cols + 2)

    ws["A3"] = "Check"
    ws["A3"].font = t["font_param_label"]
    ws["B3"] = "All cells should be 0 (Hardcoded - EPM)"
    ws["B3"].font = Font(name="Aptos", size=10, italic=True, color=DARK_RED)

    # Copy headers
    for r in [6, 7]:
        for ci in range(num_data_cols + 2):
            col_idx = ci + 1
            cl = get_column_letter(col_idx)
            c = ws.cell(row=r, column=col_idx)
            c.value = f"={hc}!{cl}{r}"
            c.font = font_header
            c.fill = t["fill_header"]

    row = 8
    section_data_rows = []

    for section_name, accounts in section_def["sections"]:
        ws.cell(row=row, column=2, value=section_name).font = t["font_section"]
        row += 1
        section_data_rows = []

        for main_acct, acct_name in accounts:
            section_data_rows.append(row)
            ws.cell(row=row, column=1, value=main_acct).font = font_normal
            ws.cell(row=row, column=2, value=acct_name).font = font_normal
            ws.cell(row=row, column=2).alignment = Alignment(indent=2)
            is_odd = len(section_data_rows) % 2 == 1
            if is_odd:
                ws.cell(row=row, column=1).fill = t["fill_alt"]
                ws.cell(row=row, column=2).fill = t["fill_alt"]
            for ci in range(num_data_cols):
                col_idx = ci + 3
                cl = get_column_letter(col_idx)
                c = ws.cell(row=row, column=col_idx)
                c.value = f"={hc}!{cl}{row}-{ep}!{cl}{row}"
                c.number_format = NUM_FMT
                c.font = font_normal
                c.border = thin_border
                if is_odd:
                    c.fill = t["fill_alt"]
            row += 1

        # Subtotal
        ws.cell(row=row, column=2, value=f"Total {section_name}").font = font_subtotal
        ws.cell(row=row, column=2).fill = t["fill_subtotal"]
        ws.cell(row=row, column=1).fill = t["fill_subtotal"]
        for ci in range(num_data_cols):
            col_idx = ci + 3
            cl = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx)
            if section_data_rows:
                parts = [f"{cl}{r}" for r in section_data_rows]
                c.value = "=" + "+".join(parts)
            else:
                c.value = 0
            c.font = font_subtotal
            c.fill = t["fill_subtotal"]
            c.number_format = NUM_FMT
            c.border = subtotal_border
        row += 1
        row += 1

    for _, formula_label, _, _ in section_def["formulas"]:
        ws.cell(row=row, column=2, value=formula_label).font = t["font_total"]
        ws.cell(row=row, column=2).fill = t["fill_total"]
        ws.cell(row=row, column=1).fill = t["fill_total"]
        for ci in range(num_data_cols):
            col_idx = ci + 3
            cl = get_column_letter(col_idx)
            c = ws.cell(row=row, column=col_idx)
            c.value = f"={hc}!{cl}{row}-{ep}!{cl}{row}"
            c.font = Font(name="Aptos", size=11, bold=True, color=RED)
            c.fill = t["fill_total"]
            c.number_format = NUM_FMT
            c.border = total_border
        row += 1
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for ci in range(num_data_cols):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14
    ws.freeze_panes = "C8"


# ── Main ─────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generic Consolidation Report")
    parser.add_argument("--group", required=True, help="Consolidation group (e.g. GROUP_CORP)")
    parser.add_argument("--year", type=int, default=None, help="Fiscal year (auto-detect if omitted)")
    parser.add_argument("--output", default="excel/Consolidation_Report.xlsx", help="Output path")
    parser.add_argument("--include-epm", action="store_true", help="Include EPM formula sheets")
    parser.add_argument("--include-delta", action="store_true", help="Include delta comparison sheets")
    args = parser.parse_args()

    print(f"Discovering config for {args.group}...")
    cfg = discover_config(args.group, args.year)
    print(f"  Group: {cfg.group}, Year: {cfg.year}, Currency: {cfg.reporting_currency}")
    for ent in cfg.entities:
        print(f"  Entity: {ent['data_area_id']} ({ent['entity_name']}) "
              f"— {ent['accounting_currency']}, {ent['ownership_pct']*100:.0f}%")
    print(f"  P&L sections: {len(cfg.pnl_sections['sections'])} "
          f"({sum(len(a) for _, a in cfg.pnl_sections['sections'])} accounts)")
    print(f"  BS sections: {len(cfg.bs_sections['sections'])} "
          f"({sum(len(a) for _, a in cfg.bs_sections['sections'])} accounts)")

    print("Fetching entity P&L data...")
    entity_pnl = fetch_entity_pnl(cfg)
    for ent in cfg.entities:
        n = len(entity_pnl.get(ent["data_area_id"], {}))
        print(f"  {ent['data_area_id']}: {n} accounts")

    print("Fetching entity BS data...")
    entity_bs = fetch_entity_bs(cfg)
    for ent in cfg.entities:
        n = len(entity_bs.get(ent["data_area_id"], {}))
        print(f"  {ent['data_area_id']}: {n} accounts")

    print("Computing Net Income plug...")
    ni_data = compute_net_income_plug(entity_bs, cfg.entities)

    print("Fetching consolidated P&L waterfall...")
    consol_pnl = fetch_consolidated_data(cfg)
    for adj_type in consol_pnl:
        n = sum(1 for a in consol_pnl[adj_type]
                for q in cfg.quarters if consol_pnl[adj_type][a].get(q, 0) != 0)
        print(f"  {adj_type}: {n} non-zero entries")

    print("Fetching consolidated BS waterfall...")
    consol_bs = fetch_consolidated_bs(cfg)

    print("Fetching NCI data...")
    pnl_nci, _ = fetch_nci_data(cfg)
    bs_nci = fetch_nci_bs_cumulative(cfg)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: P&L Entity
    print("Building P&L Entity...")
    ws = wb.create_sheet("P&L Entity")
    build_entity_sheet(ws, cfg, cfg.pnl_sections, entity_pnl, THEME_PNL,
                       sign_flip=True, is_bs=False)

    # Sheet 2: P&L Consolidated
    print("Building P&L Consolidated...")
    ws = wb.create_sheet("P&L Consolidated")
    build_consolidated_sheet(ws, cfg, cfg.pnl_sections, consol_pnl, pnl_nci,
                            THEME_CONSOL, is_bs=False, sign_flip=True)

    # Sheet 3: BS Entity
    print("Building BS Entity...")
    ws = wb.create_sheet("BS Entity")
    build_entity_sheet(ws, cfg, cfg.bs_sections, entity_bs, THEME_BS,
                       sign_flip=True, net_income=ni_data, is_bs=True)

    # Sheet 4: BS Consolidated
    print("Building BS Consolidated...")
    ws = wb.create_sheet("BS Consolidated")
    build_consolidated_sheet(ws, cfg, cfg.bs_sections, consol_bs, bs_nci,
                            THEME_CONSOL, is_bs=True, sign_flip=True)

    # Sheet 5: Bridge
    print("Building Bridge...")
    ws = wb.create_sheet("Bridge")
    build_bridge_sheet(ws, cfg, consol_pnl, consol_bs, pnl_nci, bs_nci)

    # Sheet 6: Diagnostics
    print("Building Diagnostics...")
    ws = wb.create_sheet("Diagnostics")
    build_diagnostics_sheet(ws, cfg, entity_pnl, entity_bs, consol_pnl, consol_bs)

    # Optional EPM sheets
    if args.include_epm:
        epm_pnl_periods = [("Q1", "Q1"), ("Q2", "Q2"), ("Q3", "Q3"), ("Q4", "Q4")]
        epm_bs_periods = [(3, "Q1"), (6, "Q2"), (9, "Q3"), (12, "Q4")]

        print("Building EPM sheets...")
        ws = wb.create_sheet("P&L Entity (EPM)")
        build_entity_sheet_epm(ws, cfg, cfg.pnl_sections, THEME_PNL,
                               measure="period_net_amount", periods=epm_pnl_periods,
                               sign_flip=True, is_bs=False)

        ws = wb.create_sheet("BS Entity (EPM)")
        build_entity_sheet_epm(ws, cfg, cfg.bs_sections, THEME_BS,
                               measure="ytd_net_amount", periods=epm_bs_periods,
                               sign_flip=True, is_bs=True)

    # Optional Delta sheets
    if args.include_delta and args.include_epm:
        print("Building Delta sheets...")
        n_entity_cols = len(cfg.entities) * len(cfg.quarters)

        ws = wb.create_sheet("P&L Entity (Delta)")
        build_delta_sheet(ws, "P&L Entity", "P&L Entity (EPM)",
                          cfg.pnl_sections, n_entity_cols)

        ws = wb.create_sheet("BS Entity (Delta)")
        build_delta_sheet(ws, "BS Entity", "BS Entity (EPM)",
                          cfg.bs_sections, n_entity_cols)

    # Save
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
