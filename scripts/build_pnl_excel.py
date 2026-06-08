"""Build P&L, Balance Sheet, and Cash Flow Excel workbook with =EPM() formulas."""
import sys
sys.path.insert(0, "/home/pd/frappe-bench/env/lib/python3.10/site-packages")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── P&L Structure ──────────────────────────────────────────
PNL_SECTIONS = [
    ("_HEADER", "Revenue", 0),
    ("401100", "Product Sales", 1),
    ("401200", "Service Revenue", 1),
    ("401300", "Other Revenue", 1),
    ("401400", "Accrued Sales", 1),
    ("402300", "Intercompany Sales", 1),
    ("403150", "Miscellaneous Charges", 1),
    ("403300", "Customer Cash Discounts", 1),
    ("411100", "Revenue - Labor", 1),
    ("411400", "Revenue - Fees", 1),
    ("420200", "Accrued Revenue", 1),
    ("_SUBTOTAL", "Total Revenue", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Cost of Goods Sold", 0),
    ("500100", "COGS - Finished Goods", 1),
    ("500150", "Deferred COGS", 1),
    ("510310", "Purchase Price Variance", 1),
    ("510500", "Inventory Profit/Loss", 1),
    ("510520", "Inventory Adjustment", 1),
    ("540100", "Cost of Project - Labor", 1),
    ("540110", "Cost of Project - Products", 1),
    ("6112015", "COGS - Materials & Services", 1),
    ("_SUBTOTAL", "Total COGS", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Gross Profit", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Operating Expenses", 0),
    ("600180", "Raw Materials Receipts", 1),
    ("601200", "Selling Expense", 1),
    ("601300", "Advertising Expense", 1),
    ("601500", "Travel Expense", 1),
    ("602100", "Salaries & Wages", 1),
    ("602120", "Payroll Tax", 1),
    ("602130", "Health/Life Insurance", 1),
    ("602180", "Pension/Profit-Sharing", 1),
    ("604500", "Lease Expense", 1),
    ("605150", "Rent Expense", 1),
    ("605160", "Utilities", 1),
    ("606200", "Insurance Expense", 1),
    ("606300", "Office Expense", 1),
    ("607200", "Depreciation - Tangible", 1),
    ("618900", "Miscellaneous Expense", 1),
    ("_SUBTOTAL", "Total Operating Expenses", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Operating Income", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Other Income / Expense", 0),
    ("700200", "Interest Income - Customers", 1),
    ("800200", "Partners - Interest", 1),
    ("801100", "Gain/Loss - Disposal of Assets", 1),
    ("801200", "Gain/Loss - Revaluation", 1),
    ("801400", "Currency Loss - Realized", 1),
    ("801500", "Currency Gain - Realized", 1),
    ("_SUBTOTAL", "Total Other Income/Expense", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Net Income Before Tax", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Income Tax", 0),
    ("802100", "Federal Income Tax", 1),
    ("802200", "State Income Tax", 1),
    ("_SUBTOTAL", "Total Income Tax", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Net Income", 0),
]

PNL_FORMULAS = {
    "Gross Profit": lambda st, cl: f"={cl}{st['Total Revenue']}+{cl}{st['Total COGS']}",
    "Operating Income": lambda st, cl: f"={cl}{st['Gross Profit']}-{cl}{st['Total Operating Expenses']}",
    "Net Income Before Tax": lambda st, cl: f"={cl}{st['Operating Income']}+{cl}{st['Total Other Income/Expense']}",
    "Net Income": lambda st, cl: f"={cl}{st['Net Income Before Tax']}-{cl}{st['Total Income Tax']}",
}

# ── Balance Sheet Structure ────────────────────────────────
# BS uses ytd_net_amount — cumulative position at end of period
BS_SECTIONS = [
    # --- Current Assets ---
    ("_HEADER", "Current Assets", 0),
    ("110110", "Bank Account - USD", 1),
    ("110112", "Bank Account - JPY", 1),
    ("110115", "Bank Account - CAD", 1),
    ("110125", "Bank Account - AUD", 1),
    ("110130", "Bank Account - EUR", 1),
    ("110140", "Bank Account - CHF", 1),
    ("110150", "Bank Account - GBP", 1),
    ("110160", "Bank Account - Payroll", 1),
    ("110180", "Petty Cash", 1),
    ("130100", "Accounts Receivable - Domestic", 1),
    ("130110", "Accounts Receivable - Foreign", 1),
    ("130300", "Accounts Receivable - Not Invoiced", 1),
    ("130700", "Other Receivables", 1),
    ("132190", "Prepaid Expenses", 1),
    ("133300", "Intercompany Receivable - USMF/DEMF", 1),
    ("133500", "Interunit Receivable", 1),
    ("140100", "Raw Materials Inventory", 1),
    ("140200", "Finished Goods Inventory", 1),
    ("_SUBTOTAL", "Total Current Assets", 0),

    ("_BLANK", "", 0),

    # --- Non-Current Assets ---
    ("_HEADER", "Non-Current Assets", 0),
    ("120200", "Marketable Securities", 1),
    ("134000", "Investment in Subsidiary", 1),
    ("170150", "Goodwill", 1),
    ("180100", "Tangible Fixed Assets", 1),
    ("180140", "Intangible Fixed Assets", 1),
    ("180141", "Assets Under Construction", 1),
    ("180150", "Right-of-Use Asset - Finance Leases", 1),
    ("180160", "Right-of-Use Asset - Operating Leases", 1),
    ("180200", "Accumulated Depreciation - Tangible", 1),
    ("180240", "Accumulated Amortization - Intangible", 1),
    ("180295", "Other Assets", 1),
    ("_SUBTOTAL", "Total Non-Current Assets", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Total Assets", 0),

    ("_BLANK", "", 0),

    # --- Current Liabilities ---
    ("_HEADER", "Current Liabilities", 0),
    ("200100", "Accounts Payable - Domestic", 1),
    ("200110A", "Accounts Payable - Other", 1),
    ("200130", "Invoice Pending Approval", 1),
    ("200140", "Accrued Purchases - Received Not Invoiced", 1),
    ("200190", "Accrued Purchases", 1),
    ("201100", "Salaries & Wages Payable", 1),
    ("201110", "Commissions Payable", 1),
    ("212160", "Customer Deposits", 1),
    ("221240", "Income Taxes Payable", 1),
    ("222200", "VAT Tax Payable", 1),
    ("222300", "Short Term Obligation - Finance Leases", 1),
    ("222310", "Short Term Obligation - Operating Leases", 1),
    ("231300", "Intercompany Payable - USMF/DEMF", 1),
    ("231500", "Interunit Payable", 1),
    ("_SUBTOTAL", "Total Current Liabilities", 0),

    ("_BLANK", "", 0),

    # --- Non-Current Liabilities ---
    ("_HEADER", "Non-Current Liabilities", 0),
    ("250100", "Long-term Bank Loans", 1),
    ("250200", "Notes Payable", 1),
    ("250300", "Mortgage Payable", 1),
    ("250500", "Deferred Income Tax", 1),
    ("250600", "Deferred Revenue", 1),
    ("250601", "Long Term Obligation - Finance Leases", 1),
    ("250602", "Long Term Obligation - Operating Leases", 1),
    ("_SUBTOTAL", "Total Non-Current Liabilities", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Total Liabilities", 0),

    ("_BLANK", "", 0),

    # --- Equity ---
    ("_HEADER", "Shareholders' Equity", 0),
    ("300110", "Capital Stock", 1),
    ("300120", "Paid-in Capital", 1),
    ("300130", "Dividends Paid", 1),
    ("300150", "Unrealized Currency Gain/Loss", 1),
    ("300160", "Retained Earnings", 1),
    ("300170", "Accumulated Other Comprehensive Income", 1),
    ("_SUBTOTAL", "Total Equity", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Total Liabilities & Equity", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Check (Assets - L&E)", 0),
]

BS_FORMULAS = {
    "Total Assets": lambda st, cl: f"={cl}{st['Total Current Assets']}+{cl}{st['Total Non-Current Assets']}",
    "Total Liabilities": lambda st, cl: f"={cl}{st['Total Current Liabilities']}+{cl}{st['Total Non-Current Liabilities']}",
    "Total Liabilities & Equity": lambda st, cl: f"={cl}{st['Total Liabilities']}+{cl}{st['Total Equity']}",
    "Check (Assets - L&E)": lambda st, cl: f"={cl}{st['Total Assets']}-{cl}{st['Total Liabilities & Equity']}",
}

# ── Cash Flow Structure ────────────────────────────────────
# Indirect method: start from net income, adjust for non-cash items
CF_SECTIONS = [
    ("_HEADER", "Operating Activities", 0),
    # Net income from P&L (manual reference — we use a _LINK tag)
    ("_LINK_NI", "Net Income", 1),
    ("_SPACER", "Adjustments for non-cash items:", 0),
    ("607200", "Depreciation - Tangible Assets", 1),
    ("607100", "Depreciation - Intangible Assets", 1),
    ("801100", "Gain/Loss - Disposal of Assets", 1),
    ("801200", "Gain/Loss - Revaluation", 1),
    ("_SPACER", "Changes in working capital:", 0),
    ("130100", "Accounts Receivable", 1),
    ("130300", "Accounts Receivable - Not Invoiced", 1),
    ("140100", "Raw Materials Inventory", 1),
    ("140200", "Finished Goods Inventory", 1),
    ("132190", "Prepaid Expenses", 1),
    ("200100", "Accounts Payable", 1),
    ("200140", "Accrued Purchases", 1),
    ("200190", "Accrued Purchases - Other", 1),
    ("_SUBTOTAL", "Net Cash from Operating Activities", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Investing Activities", 0),
    ("180100", "Purchase of Fixed Assets", 1),
    ("180141", "Assets Under Construction", 1),
    ("180140", "Purchase of Intangible Assets", 1),
    ("120200", "Marketable Securities", 1),
    ("_SUBTOTAL", "Net Cash from Investing Activities", 0),

    ("_BLANK", "", 0),

    ("_HEADER", "Financing Activities", 0),
    ("250100", "Long-term Bank Loans", 1),
    ("250200", "Notes Payable", 1),
    ("250300", "Mortgage Payable", 1),
    ("300110", "Capital Stock", 1),
    ("300130", "Dividends Paid", 1),
    ("_SUBTOTAL", "Net Cash from Financing Activities", 0),

    ("_BLANK", "", 0),
    ("_FORMULA", "Net Change in Cash", 0),

    ("_BLANK", "", 0),

    # Opening / Closing cash (YTD-based)
    ("_HEADER", "Cash Position", 0),
    ("_CASH_OPEN", "Opening Cash (prior year-end)", 1),
    ("_FORMULA_REF", "Net Change in Cash (above)", 1),
    ("_CASH_CLOSE", "Closing Cash", 1),
]

CF_FORMULAS = {
    "Net Change in Cash": lambda st, cl: (
        f"={cl}{st['Net Cash from Operating Activities']}"
        f"+{cl}{st['Net Cash from Investing Activities']}"
        f"+{cl}{st['Net Cash from Financing Activities']}"
    ),
}

# Cash accounts to sum for opening/closing balance
CASH_ACCOUNTS = ["110110", "110112", "110115", "110125", "110130",
                  "110140", "110150", "110160", "110180"]

ENTITIES = ["USMF", "DEMF"]
YEARS = [2024, 2025]

# P&L periods (monthly + aggregates)
PNL_PERIODS = [
    (1, "Jan"), (2, "Feb"), (3, "Mar"), ("Q1", "Q1"),
    (4, "Apr"), (5, "May"), (6, "Jun"), ("Q2", "Q2"),
    (7, "Jul"), (8, "Aug"), (9, "Sep"), ("Q3", "Q3"),
    (10, "Oct"), (11, "Nov"), (12, "Dec"), ("Q4", "Q4"),
    ("H1", "H1"), ("H2", "H2"), ("FY", "Full Year"),
]

# BS periods (quarterly snapshots — YTD at end of Q)
BS_PERIODS = [
    (3, "Q1"), (6, "Q2"), (9, "Q3"), (12, "Q4"),
]

# CF periods (quarterly movements)
CF_PERIODS = [
    ("Q1", "Q1"), ("Q2", "Q2"), ("Q3", "Q3"), ("Q4", "Q4"),
    ("H1", "H1"), ("H2", "H2"), ("FY", "Full Year"),
]

# ── Styles ──────────────────────────────────────────────────
DARK_BLUE = "1F3864"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
DARK_GREEN = "1D4E2E"
MED_GREEN = "2E7D46"
LIGHT_GREEN = "D6EFD6"
DARK_TEAL = "1B4D5C"
MED_TEAL = "2A7A8C"
LIGHT_TEAL = "D0ECF0"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "C00000"

font_title = Font(name="Aptos", size=14, bold=True, color=WHITE)
font_header = Font(name="Aptos", size=10, bold=True, color=WHITE)
font_section = Font(name="Aptos", size=10, bold=True)
font_subtotal = Font(name="Aptos", size=10, bold=True)
font_total = Font(name="Aptos", size=11, bold=True)
font_normal = Font(name="Aptos", size=10)
font_param_label = Font(name="Aptos", size=10, bold=True)
font_param_value = Font(name="Aptos", size=10, bold=True)
font_spacer = Font(name="Aptos", size=9, italic=True, color="666666")

thin_border = Border(bottom=Side(style="thin", color="B0B0B0"))
subtotal_border = Border(top=Side(style="thin", color=BLACK), bottom=Side(style="thin", color=BLACK))
total_border = Border(top=Side(style="thin", color=BLACK), bottom=Side(style="double", color=BLACK))

NUM_FMT = '#,##0'


def _make_theme(primary_dark, primary_med, primary_light):
    """Return a style dict for a given color theme."""
    return {
        "fill_title": PatternFill(start_color=primary_dark, end_color=primary_dark, fill_type="solid"),
        "fill_header": PatternFill(start_color=primary_med, end_color=primary_med, fill_type="solid"),
        "fill_subtotal": PatternFill(start_color=primary_light, end_color=primary_light, fill_type="solid"),
        "fill_total": PatternFill(start_color=primary_light, end_color=primary_light, fill_type="solid"),
        "fill_alt": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
        "font_section": Font(name="Aptos", size=10, bold=True, color=primary_dark),
        "font_total": Font(name="Aptos", size=11, bold=True, color=primary_dark),
        "font_param_label": Font(name="Aptos", size=10, bold=True, color=primary_dark),
    }


THEME_PNL = _make_theme(DARK_BLUE, MED_BLUE, LIGHT_BLUE)
THEME_BS = _make_theme(DARK_GREEN, MED_GREEN, LIGHT_GREEN)
THEME_CF = _make_theme(DARK_TEAL, MED_TEAL, LIGHT_TEAL)


def _build_generic_sheet(ws, entity, year, title, sections, formulas,
                         periods, measure, sign_flip, theme,
                         extra_builder=None):
    """Generic sheet builder for P&L, BS, and CF."""
    t = theme
    entity_ref = "$B$3"
    year_ref = "$B$4"

    # ── Title row ───────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].fill = t["fill_title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, len(periods) + 3):
        ws.cell(row=1, column=col).fill = t["fill_title"]

    # ── Parameters ──────────────────────────────────────────
    ws["A3"] = "Entity"
    ws["A3"].font = t["font_param_label"]
    ws["B3"] = entity
    ws["B3"].font = font_param_value

    ws["A4"] = "Year"
    ws["A4"].font = t["font_param_label"]
    ws["B4"] = year
    ws["B4"].font = font_param_value

    ws["D3"] = "Measure"
    ws["D3"].font = t["font_param_label"]
    ws["E3"] = measure
    ws["E3"].font = font_param_value

    # ── Column headers ──────────────────────────────────────
    header_row = 6
    ws.cell(row=header_row, column=1, value="Account").font = font_header
    ws.cell(row=header_row, column=1).fill = t["fill_header"]
    ws.cell(row=header_row, column=2, value="Description").font = font_header
    ws.cell(row=header_row, column=2).fill = t["fill_header"]

    for pi, (pval, plabel) in enumerate(periods):
        c = ws.cell(row=header_row, column=pi + 3, value=plabel)
        c.font = font_header
        c.fill = t["fill_header"]
        c.alignment = Alignment(horizontal="center")

    # ── Data rows ───────────────────────────────────────────
    row = 7
    subtotal_rows = {}
    formula_queue = []
    section_data_rows = []
    sign = "-" if sign_flip else ""

    for item in sections:
        tag, label, indent = item

        if tag == "_BLANK":
            row += 1
            continue

        if tag == "_SPACER":
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = font_spacer
            row += 1
            continue

        if tag == "_HEADER":
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = t["font_section"]
            section_data_rows = []
            row += 1
            continue

        if tag == "_SUBTOTAL":
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = font_subtotal
            ws.cell(row=row, column=2).fill = t["fill_subtotal"]
            ws.cell(row=row, column=1).fill = t["fill_subtotal"]

            for pi in range(len(periods)):
                col = pi + 3
                c = ws.cell(row=row, column=col)
                if section_data_rows:
                    parts = [f"{get_column_letter(col)}{r}" for r in section_data_rows]
                    c.value = "=" + "+".join(parts)
                else:
                    c.value = 0
                c.font = font_subtotal
                c.fill = t["fill_subtotal"]
                c.number_format = NUM_FMT
                c.border = subtotal_border

            subtotal_rows[label] = row
            row += 1
            continue

        if tag == "_FORMULA":
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = t["font_total"]
            ws.cell(row=row, column=2).fill = t["fill_total"]
            ws.cell(row=row, column=1).fill = t["fill_total"]
            formula_queue.append((row, label))
            subtotal_rows[label] = row

            is_check = "Check" in label
            for pi in range(len(periods)):
                col = pi + 3
                c = ws.cell(row=row, column=col)
                c.font = Font(name="Aptos", size=11, bold=True, color=RED) if is_check else t["font_total"]
                c.fill = t["fill_total"]
                c.number_format = NUM_FMT
                c.border = total_border if "Total" in label or "Net Income" == label or is_check else subtotal_border

            row += 1
            continue

        # Special CF tags handled by extra_builder
        if tag.startswith("_") and extra_builder:
            extra_row = extra_builder(ws, row, tag, label, periods, entity_ref,
                                      year_ref, subtotal_rows, t, section_data_rows)
            if extra_row:
                row = extra_row
            continue

        # ── Data row ────────────────────────────────────────
        acct_id = tag
        section_data_rows.append(row)

        ws.cell(row=row, column=1).value = acct_id
        ws.cell(row=row, column=1).font = font_normal
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = font_normal
        ws.cell(row=row, column=2).alignment = Alignment(indent=2)

        is_odd = len(section_data_rows) % 2 == 1
        if is_odd:
            ws.cell(row=row, column=1).fill = t["fill_alt"]
            ws.cell(row=row, column=2).fill = t["fill_alt"]

        for pi, (pval, plabel) in enumerate(periods):
            col = pi + 3
            c = ws.cell(row=row, column=col)
            parg = f'"{pval}"' if isinstance(pval, str) else str(pval)
            c.value = f'={sign}EPM({entity_ref},{year_ref},{parg},$A{row},"{measure}")'
            c.number_format = NUM_FMT
            c.font = font_normal
            c.border = thin_border
            if is_odd:
                c.fill = t["fill_alt"]

        row += 1

    # ── Fill formulas ───────────────────────────────────────
    for frow, flabel in formula_queue:
        if flabel in formulas:
            fn = formulas[flabel]
            for pi in range(len(periods)):
                col = pi + 3
                cl = get_column_letter(col)
                ws.cell(row=frow, column=col).value = fn(subtotal_rows, cl)

    # ── Column widths ───────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for pi in range(len(periods)):
        cl = get_column_letter(pi + 3)
        pval = periods[pi][0]
        ws.column_dimensions[cl].width = 14 if isinstance(pval, str) else 13

    ws.freeze_panes = "C7"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _cf_extra_builder(ws, row, tag, label, periods, entity_ref, year_ref,
                      subtotal_rows, t, section_data_rows):
    """Handle special Cash Flow tags."""
    if tag == "_LINK_NI":
        # Net Income line — uses P&L measure period_net_amount, sign-flipped
        section_data_rows.append(row)
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = font_normal
        ws.cell(row=row, column=2).alignment = Alignment(indent=2)

        for pi, (pval, plabel) in enumerate(periods):
            col = pi + 3
            c = ws.cell(row=row, column=col)
            parg = f'"{pval}"' if isinstance(pval, str) else str(pval)
            # Sum of all P&L accounts (Revenue is negative in GL, expenses positive)
            # Net income = -(sum of all P&L) = flip sign
            # Use a simple placeholder: sum key revenue/expense accounts
            # Actually, just reference the retained earnings movement as a proxy,
            # or we hardcode the P&L total. Simplest: pull from retained earnings account.
            c.value = f'=-EPM({entity_ref},{year_ref},{parg},"300160","period_net_amount")'
            c.number_format = NUM_FMT
            c.font = font_normal
            c.border = thin_border

        return row + 1

    if tag == "_CASH_OPEN":
        # Opening cash: YTD of cash accounts at prior year-end (period 12 of year-1)
        section_data_rows.append(row)
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = font_normal
        ws.cell(row=row, column=2).alignment = Alignment(indent=2)

        for pi in range(len(periods)):
            col = pi + 3
            c = ws.cell(row=row, column=col)
            # Sum ytd_net_amount at P12 of prior year for all cash accounts
            parts = []
            for ca in CASH_ACCOUNTS:
                parts.append(f'EPM({entity_ref},{year_ref}-1,12,"{ca}","ytd_net_amount")')
            c.value = "=-(" + "+".join(parts) + ")"
            c.number_format = NUM_FMT
            c.font = font_normal
            c.border = thin_border

        return row + 1

    if tag == "_FORMULA_REF":
        # Reference to Net Change in Cash (already computed above)
        section_data_rows.append(row)
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = font_normal
        ws.cell(row=row, column=2).alignment = Alignment(indent=2)

        ncc_row = subtotal_rows.get("Net Change in Cash")
        if ncc_row:
            for pi in range(len(periods)):
                col = pi + 3
                cl = get_column_letter(col)
                c = ws.cell(row=row, column=col)
                c.value = f"={cl}{ncc_row}"
                c.number_format = NUM_FMT
                c.font = font_normal
                c.border = thin_border

        return row + 1

    if tag == "_CASH_CLOSE":
        # Closing = Opening + Net Change
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = t["font_total"]
        ws.cell(row=row, column=2).fill = t["fill_total"]
        ws.cell(row=row, column=1).fill = t["fill_total"]

        # Opening is 2 rows back, Net Change is 1 row back
        open_row = row - 2
        change_row = row - 1
        for pi in range(len(periods)):
            col = pi + 3
            cl = get_column_letter(col)
            c = ws.cell(row=row, column=col)
            c.value = f"={cl}{open_row}+{cl}{change_row}"
            c.number_format = NUM_FMT
            c.font = t["font_total"]
            c.fill = t["fill_total"]
            c.border = total_border

        return row + 1

    return None


def main():
    wb = Workbook()
    wb.remove(wb.active)

    for entity in ENTITIES:
        for year in YEARS:
            # P&L sheet
            ws = wb.create_sheet(title=f"P&L {entity} {year}")
            _build_generic_sheet(
                ws, entity, year,
                title="Profit & Loss Statement",
                sections=PNL_SECTIONS,
                formulas=PNL_FORMULAS,
                periods=PNL_PERIODS,
                measure="period_net_amount",
                sign_flip=True,
                theme=THEME_PNL,
            )

            # Balance Sheet
            ws = wb.create_sheet(title=f"BS {entity} {year}")
            _build_generic_sheet(
                ws, entity, year,
                title="Balance Sheet",
                sections=BS_SECTIONS,
                formulas=BS_FORMULAS,
                periods=BS_PERIODS,
                measure="ytd_net_amount",
                sign_flip=True,
                theme=THEME_BS,
            )

            # Cash Flow
            ws = wb.create_sheet(title=f"CF {entity} {year}")
            _build_generic_sheet(
                ws, entity, year,
                title="Statement of Cash Flows",
                sections=CF_SECTIONS,
                formulas=CF_FORMULAS,
                periods=CF_PERIODS,
                measure="period_net_amount",
                sign_flip=False,
                theme=THEME_CF,
                extra_builder=_cf_extra_builder,
            )

    out_path = "/home/pd/open_epm/excel/Open_EPM_PnL.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
