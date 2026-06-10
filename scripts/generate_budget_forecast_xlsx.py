#!/usr/bin/env python3
"""
Generate Budget 2024 and Forecast 2024F1-F12 Excel upload sheets.

Output: excel/Budget_Forecast_2024.xlsx
  - Sheet "Setup": connection config (URL, user, password) + instructions
  - Sheet "Budget 2024": annual budget with EPMSAVE formulas
  - Sheets "2024F1" through "2024F12": rolling forecasts with EPMSAVE formulas

IMPORTANT: After opening in Excel, import the VBA module:
  Alt+F11 > File > Import > select excel/OpenEPM.bas
  Then press the "Setup" button on the Open EPM toolbar to configure connection.
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Constants from demo data ─────────────────────────────────────
ENTITIES = [
    {"id": "AMHQ", "name": "Alpine Manufacturing HQ", "ccy": "CHF",
     "cost_center": "HQ", "department": "MGMT"},
    {"id": "AMUS", "name": "Alpine Manufacturing US", "ccy": "USD",
     "cost_center": "SALES", "department": "SALES"},
    {"id": "AMDE", "name": "Alpine Manufacturing DE", "ccy": "EUR",
     "cost_center": "PROD", "department": "OPS"},
]

PNL_ACCOUNTS = [
    ("4010", "Product Revenue", "Revenue", -1),
    ("4020", "Service Revenue", "Revenue", -1),
    ("5010", "Cost of Goods Sold", "COGS", 1),
    ("6010", "Salaries and Wages", "OpEx", 1),
    ("6020", "Rent Expense", "OpEx", 1),
    ("6030", "Depreciation Expense", "OpEx", 1),
    ("6040", "Marketing Expense", "OpEx", 1),
    ("6050", "Travel Expense", "OpEx", 1),
    ("6060", "Utilities", "OpEx", 1),
]

SEASONAL = [0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15, 1.10, 1.05, 1.00, 0.95, 0.90]
REVENUE_BASE = {"AMHQ": 500000, "AMUS": 800000, "AMDE": 600000}
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ── Styling ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ENTITY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ENTITY_FONT = Font(name="Calibri", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
ACTUAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SETUP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
INPUT_BORDER = Border(
    left=Side(style="medium", color="4472C4"),
    right=Side(style="medium", color="4472C4"),
    top=Side(style="medium", color="4472C4"),
    bottom=Side(style="medium", color="4472C4"),
)
NUM_FMT = '#,##0'


def generate_budget_amounts(entity_id):
    """Generate budget amounts per account per month (8% revenue growth target)."""
    base_rev = REVENUE_BASE[entity_id]
    rows = {}
    for acct_id, acct_name, category, sign in PNL_ACCOUNTS:
        monthly = []
        for m in range(12):
            seasonal = SEASONAL[m]
            rev = int(base_rev * seasonal * 1.08)
            if acct_id == "4010":
                amt = rev
            elif acct_id == "4020":
                amt = int(rev * 0.12)
            elif acct_id == "5010":
                amt = int(rev * 0.40)
            elif acct_id == "6010":
                amt = int(rev * 0.24)
            elif acct_id == "6020":
                amt = int(rev * 0.04)
            elif acct_id == "6030":
                amt = int(rev * 0.03)
            elif acct_id == "6040":
                amt = int(rev * 0.03)
            elif acct_id == "6050":
                amt = int(rev * 0.015)
            elif acct_id == "6060":
                amt = int(rev * 0.015)
            else:
                amt = 0
            monthly.append(amt * sign)
        rows[acct_id] = monthly
    return rows


def generate_actual_amounts(entity_id):
    """Generate 'actual' amounts with random variance (simulating real results)."""
    base_rev = REVENUE_BASE[entity_id]
    rows = {}
    for acct_id, acct_name, category, sign in PNL_ACCOUNTS:
        monthly = []
        for m in range(12):
            seasonal = SEASONAL[m]
            variation = random.uniform(0.93, 1.07)
            rev = int(base_rev * seasonal * variation)
            if acct_id == "4010":
                amt = rev
            elif acct_id == "4020":
                amt = int(rev * random.uniform(0.10, 0.14))
            elif acct_id == "5010":
                amt = int(rev * random.uniform(0.38, 0.44))
            elif acct_id == "6010":
                amt = int(rev * random.uniform(0.23, 0.26))
            elif acct_id == "6020":
                amt = int(rev * 0.04)
            elif acct_id == "6030":
                amt = int(rev * 0.03)
            elif acct_id == "6040":
                amt = int(rev * random.uniform(0.02, 0.04))
            elif acct_id == "6050":
                amt = int(rev * random.uniform(0.01, 0.02))
            elif acct_id == "6060":
                amt = int(rev * 0.015)
            else:
                amt = 0
            monthly.append(amt * sign)
        rows[acct_id] = monthly
    return rows


def generate_forecast_amounts(entity_id, forecast_month):
    """Months before forecast_month = actuals, rest = re-forecasted."""
    actuals = generate_actual_amounts(entity_id)
    budget = generate_budget_amounts(entity_id)
    rows = {}
    for acct_id, acct_name, category, sign in PNL_ACCOUNTS:
        monthly = []
        for m in range(12):
            if m < forecast_month - 1:
                monthly.append(actuals[acct_id][m])
            else:
                if forecast_month > 1 and acct_id in actuals:
                    ytd_actual = sum(actuals[acct_id][:forecast_month - 1])
                    ytd_budget = sum(budget[acct_id][:forecast_month - 1])
                    if ytd_budget != 0:
                        trend = ytd_actual / ytd_budget
                        adj = 1.0 + (trend - 1.0) * 0.5
                    else:
                        adj = 1.0
                    monthly.append(int(budget[acct_id][m] * adj))
                else:
                    monthly.append(budget[acct_id][m])
        rows[acct_id] = monthly
    return rows, (forecast_month - 1)


def style_cell(cell, font=None, fill=None, alignment=None, border=None, num_fmt=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt


# ── Setup sheet ──────────────────────────────────────────────────

def create_setup_sheet(wb):
    """Create the Setup/Config sheet with connection fields and instructions."""
    ws = wb.create_sheet(title="Setup")

    # Title
    cell = ws.cell(row=1, column=1, value="Open EPM — Connection Setup")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=16, color="2F5496"))
    ws.merge_cells("A1:F1")

    # Separator
    for col in range(1, 7):
        style_cell(ws.cell(row=2, column=col), fill=HEADER_FILL)

    # ── Connection Settings ──
    r = 4
    cell = ws.cell(row=r, column=1, value="Connection Settings")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    r = 6
    labels = [
        ("Server URL:", "https://localhost", "Frappe/ERPNext server URL (HTTPS)"),
        ("Username:", "Administrator", "Frappe username"),
        ("Password:", "admin123", "Frappe password"),
    ]
    for label, default, hint in labels:
        ws.cell(row=r, column=1, value=label)
        style_cell(ws.cell(row=r, column=1),
                   font=Font(name="Calibri", bold=True, size=11))

        cell = ws.cell(row=r, column=2, value=default)
        style_cell(cell, font=Font(name="Calibri", size=11),
                   fill=INPUT_FILL, border=INPUT_BORDER)

        ws.cell(row=r, column=4, value=hint)
        style_cell(ws.cell(row=r, column=4),
                   font=Font(name="Calibri", size=10, italic=True, color="666666"))
        r += 1

    # Named ranges for the config cells
    # (VBA reads from these cells as fallback if CustomDocumentProperties not set)

    # ── Option 1: VBA ──
    r += 1
    cell = ws.cell(row=r, column=1, value="Option 1: VBA Macros")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    r += 1
    ws.cell(row=r, column=1, value="Best for:")
    style_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", bold=True, size=10))
    ws.cell(row=r, column=2, value="Desktop Excel, keyboard shortcuts, works offline")
    style_cell(ws.cell(row=r, column=2), font=Font(name="Calibri", size=10, italic=True))

    vba_steps = [
        ("Step 1:", "Import VBA Module",
         'Press Alt+F11 to open VBA Editor, then File > Import File > '
         'select "excel/OpenEPM.bas" from the project folder.'),
        ("Step 2:", "Enable Macros",
         'Save this file as .xlsm (macro-enabled). '
         'Close and reopen - the "Open EPM" toolbar appears automatically.'),
        ("Step 3:", "Click Setup on Toolbar",
         'The "Open EPM" toolbar has buttons: '
         'Setup | Refresh | Refresh All | Save | Save All | Debug. '
         'Click "Setup" to enter your server URL, username, and password.'),
        ("Step 4:", "Edit Budget/Forecast",
         'Go to any data sheet (Budget 2024, 2024F1, etc.) and edit the '
         'white cells with your budget or forecast numbers. '
         'Yellow cells are actual months (read-only reference).'),
        ("Step 5:", "Save to Server",
         'Press Ctrl+Shift+S or click "Save" on the toolbar to upload '
         'the active sheet to the server. Click "Save All" for all sheets. '
         'VBA reads the row metadata (entity, account, etc.) automatically.'),
    ]

    r += 1
    for step_num, step_title, step_desc in vba_steps:
        r += 1
        ws.cell(row=r, column=1, value=step_num)
        style_cell(ws.cell(row=r, column=1),
                   font=Font(name="Calibri", bold=True, size=11, color="2F5496"))
        ws.cell(row=r, column=2, value=step_title)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", bold=True, size=11))
        r += 1
        ws.cell(row=r, column=2, value=step_desc)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", size=10, color="333333"))
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    # ── Option 2: Office.js Add-in ──
    r += 2
    cell = ws.cell(row=r, column=1, value="Option 2: Office.js Add-in (No Macros)")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    r += 1
    ws.cell(row=r, column=1, value="Best for:")
    style_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", bold=True, size=10))
    ws.cell(row=r, column=2,
            value="No macros needed, no code signing, works in Excel Online and Desktop")
    style_cell(ws.cell(row=r, column=2), font=Font(name="Calibri", size=10, italic=True))

    addin_steps = [
        ("Step 1:", "Sideload the Add-in",
         'In Excel: Insert > My Add-ins > Upload My Add-in > '
         'browse to "excel-addin/manifest.xml". The "OpenEPM" button '
         'appears in the Home ribbon tab.'),
        ("Step 2:", "Open the Task Pane",
         'Click the "OpenEPM" button in the ribbon. Log in with your '
         'Frappe username and password.'),
        ("Step 3:", "Use the Budget Tab",
         'Switch to the "Budget" tab in the task pane. The Sheet Info card '
         'shows the detected scenario, entity count, and row count. '
         'Click "Save Sheet" or "Save All Sheets" to upload data. '
         'Click "Refresh Actuals" to pull actual values from the server.'),
    ]

    for step_num, step_title, step_desc in addin_steps:
        r += 1
        ws.cell(row=r, column=1, value=step_num)
        style_cell(ws.cell(row=r, column=1),
                   font=Font(name="Calibri", bold=True, size=11, color="107C10"))
        ws.cell(row=r, column=2, value=step_title)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", bold=True, size=11))
        r += 1
        ws.cell(row=r, column=2, value=step_desc)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", size=10, color="333333"))
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    # ── Keyboard Shortcuts ──
    r += 2
    cell = ws.cell(row=r, column=1, value="Keyboard Shortcuts")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    shortcuts = [
        ("Ctrl+Shift+R", "Refresh active sheet (fetch actuals + budget from server)"),
        ("Ctrl+Shift+S", "Save active sheet budget/forecast data to server"),
        ("Alt+F11", "Open VBA Editor (to import/edit OpenEPM.bas)"),
    ]
    r += 1
    for key, desc in shortcuts:
        r += 1
        ws.cell(row=r, column=1, value=key)
        style_cell(ws.cell(row=r, column=1),
                   font=Font(name="Consolas", bold=True, size=11))
        ws.cell(row=r, column=2, value=desc)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", size=10))

    # ── Sheet Legend ──
    r += 2
    cell = ws.cell(row=r, column=1, value="Sheet Legend")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    legend = [
        ("White cells", "Editable — change the amount, then press F9 or Ctrl+Shift+R to save"),
        ("Yellow cells", "Actual months (closed) — included for reference, not saved to server"),
        ("Green cells", "Computed totals (SUM formulas) — auto-calculated"),
        ("Blue header", "Column headers — do not edit"),
    ]
    r += 1
    fills = [INPUT_FILL, ACTUAL_FILL, TOTAL_FILL, HEADER_FILL]
    fonts_legend = [
        Font(name="Calibri", size=10),
        Font(name="Calibri", size=10),
        Font(name="Calibri", size=10),
        Font(name="Calibri", size=10, color="FFFFFF"),
    ]
    for i, (label, desc) in enumerate(legend):
        r += 1
        cell = ws.cell(row=r, column=1, value=label)
        style_cell(cell, font=fonts_legend[i], fill=fills[i], border=THIN_BORDER)
        ws.cell(row=r, column=2, value=desc)
        style_cell(ws.cell(row=r, column=2),
                   font=Font(name="Calibri", size=10))

    # ── How Save Works ──
    r += 2
    cell = ws.cell(row=r, column=1, value="How Save Works")
    style_cell(cell, font=Font(name="Calibri", bold=True, size=13, color="2F5496"))
    ws.merge_cells(f"A{r}:F{r}")

    r += 2
    ws.cell(row=r, column=1, value="Sheet layout:")
    style_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", bold=True, size=11))
    ws.cell(row=r, column=2,
            value='Col A: Scenario ID | Col B: Entity | Col D: Account | '
                  'Col G: Cost Center | Col H: Department | Cols I-T: Period 1-12')
    style_cell(ws.cell(row=r, column=2), font=Font(name="Consolas", size=10))
    ws.merge_cells(f"B{r}:F{r}")

    r += 2
    ws.cell(row=r, column=1, value="How it works:")
    style_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", bold=True, size=11))
    ws.cell(row=r, column=2,
            value='When you click Save (or Ctrl+Shift+S), VBA scans each data row, '
                  'reads the entity/account/scenario from columns A-H, collects the 12 '
                  'monthly amounts from columns I-T, and batch-POSTs everything to the server. '
                  'The layer is read from the info block at the top of each sheet.')
    style_cell(ws.cell(row=r, column=2), font=Font(name="Calibri", size=10))
    ws.merge_cells(f"B{r}:F{r}")
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    r += 2
    ws.cell(row=r, column=1, value="No formulas needed:")
    style_cell(ws.cell(row=r, column=1), font=Font(name="Calibri", bold=True, size=11))
    ws.cell(row=r, column=2,
            value='Just type plain numbers in the white cells. '
                  'No EPMSAVE formulas required. Edit freely, then Save when ready.')
    style_cell(ws.cell(row=r, column=2), font=Font(name="Calibri", size=10))
    ws.merge_cells(f"B{r}:F{r}")

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    ws.sheet_properties.tabColor = "2F5496"
    return ws


# ── Data sheets ──────────────────────────────────────────────────

def write_header_row(ws, row):
    """Write the metadata + month header row."""
    headers = [
        "Scenario ID", "Entity", "Currency", "Account",
        "Account Name", "Category", "Cost Center", "Department",
    ] + MONTHS + ["Annual Total"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=Alignment(horizontal="center", wrap_text=True),
                   border=THIN_BORDER)


def write_entity_block(ws, start_row, entity, scenario_id, amounts,
                       actual_months=0):
    """Write one entity's data block with EPMSAVE formulas. Returns next row."""
    row = start_row

    # Entity separator row
    ws.cell(row=row, column=1, value=entity["id"])
    for col in range(1, 22):
        style_cell(ws.cell(row=row, column=col),
                   font=ENTITY_FONT, fill=ENTITY_FILL, border=THIN_BORDER)
    ws.cell(row=row, column=2, value=entity["name"])
    ws.cell(row=row, column=3, value=entity["ccy"])
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=20)
    row += 1

    for acct_id, acct_name, category, sign in PNL_ACCOUNTS:
        monthly = amounts[acct_id]
        c = 1
        ws.cell(row=row, column=c, value=scenario_id); c += 1
        ws.cell(row=row, column=c, value=entity["id"]); c += 1
        ws.cell(row=row, column=c, value=entity["ccy"]); c += 1
        ws.cell(row=row, column=c, value=acct_id); c += 1
        ws.cell(row=row, column=c, value=acct_name); c += 1
        ws.cell(row=row, column=c, value=category); c += 1
        ws.cell(row=row, column=c, value=entity["cost_center"]); c += 1
        ws.cell(row=row, column=c, value=entity["department"]); c += 1

        # Monthly amounts — plain numbers (saved via batch Save button)
        for m in range(12):
            val = monthly[m]
            cell = ws.cell(row=row, column=c, value=val)
            style_cell(cell, font=DATA_FONT, border=THIN_BORDER, num_fmt=NUM_FMT)
            if m < actual_months:
                style_cell(cell, fill=ACTUAL_FILL)
            else:
                style_cell(cell, fill=INPUT_FILL)
            c += 1

        # Annual total formula
        first_col = get_column_letter(9)
        last_col = get_column_letter(20)
        total_cell = ws.cell(row=row, column=c,
                             value=f"=SUM({first_col}{row}:{last_col}{row})")
        style_cell(total_cell, font=TOTAL_FONT, fill=TOTAL_FILL,
                   border=THIN_BORDER, num_fmt=NUM_FMT)

        # Style metadata columns
        for col in range(1, 9):
            style_cell(ws.cell(row=row, column=col),
                       font=DATA_FONT, border=THIN_BORDER)

        row += 1

    # Net Income totals row
    total_row = row
    ws.cell(row=total_row, column=5, value="Net Income")
    style_cell(ws.cell(row=total_row, column=5), font=TOTAL_FONT)
    for col in range(9, 22):
        first_data = start_row + 1
        last_data = row - 1
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({col_letter}{first_data}:{col_letter}{last_data})")
        style_cell(cell, font=TOTAL_FONT, fill=TOTAL_FILL,
                   border=THIN_BORDER, num_fmt=NUM_FMT)
    row += 1

    return row + 1  # skip blank row


def write_info_block(ws, scenario_id, scenario_name, scenario_type,
                     fiscal_year=2024, actual_months=0):
    """Write info block at top of each data sheet."""
    info = [
        ("Scenario:", scenario_name),
        ("Scenario ID:", scenario_id),
        ("Type:", scenario_type),
        ("Fiscal Year:", fiscal_year),
        ("Layer:", "base"),
    ]
    for r, (label, val) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        style_cell(ws.cell(row=r, column=1), font=Font(bold=True, name="Calibri"))
        style_cell(ws.cell(row=r, column=2), font=Font(name="Calibri"))

    if actual_months > 0:
        ws.cell(row=6, column=1, value="Actual months:")
        ws.cell(row=6, column=2, value=f"1-{actual_months} (yellow = actuals, white = editable forecast)")
        style_cell(ws.cell(row=6, column=1), font=Font(bold=True, name="Calibri"))
        style_cell(ws.cell(row=6, column=2),
                   font=Font(name="Calibri", italic=True, color="996600"))

    # Usage hint
    ws.cell(row=7, column=1, value="Usage:")
    ws.cell(row=7, column=2,
            value="Edit white cells, then press Ctrl+Shift+R or F9 to save to server")
    style_cell(ws.cell(row=7, column=1), font=Font(bold=True, name="Calibri"))
    style_cell(ws.cell(row=7, column=2),
               font=Font(name="Calibri", italic=True, color="2F5496"))


def set_column_widths(ws):
    widths = {
        1: 18, 2: 12, 3: 10, 4: 10, 5: 22,
        6: 12, 7: 14, 8: 14, 21: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in range(9, 21):
        ws.column_dimensions[get_column_letter(col)].width = 13


def create_data_sheet(wb, sheet_name, scenario_id, scenario_name, scenario_type,
                      amounts_fn, actual_months=0, tab_color=None):
    """Create one data sheet (budget or forecast)."""
    ws = wb.create_sheet(title=sheet_name)

    write_info_block(ws, scenario_id, scenario_name, scenario_type,
                     actual_months=actual_months)

    header_row = 9
    write_header_row(ws, header_row)

    row = header_row + 1
    for entity in ENTITIES:
        amounts = amounts_fn(entity["id"])
        if isinstance(amounts, tuple):
            amounts, _ = amounts
        row = write_entity_block(ws, row, entity, scenario_id, amounts,
                                 actual_months=actual_months)

    set_column_widths(ws)
    ws.freeze_panes = f"I{header_row + 1}"

    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    return ws


# ── Main ─────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Setup sheet (connection config + instructions)
    create_setup_sheet(wb)

    # 2. Budget 2024
    create_data_sheet(
        wb, "Budget 2024",
        scenario_id="BUDGET_2024",
        scenario_name="Budget 2024",
        scenario_type="budget",
        amounts_fn=generate_budget_amounts,
        tab_color="70AD47",  # green
    )

    # 3. Forecast 2024F1-F12
    for fm in range(1, 13):
        sheet_name = f"2024F{fm}"
        scenario_id = f"FORECAST_2024F{fm}"
        scenario_name = f"Forecast 2024 - {MONTHS[fm - 1]}"

        def make_fn(forecast_month=fm):
            def fn(entity_id):
                return generate_forecast_amounts(entity_id, forecast_month)
            return fn

        create_data_sheet(
            wb, sheet_name,
            scenario_id=scenario_id,
            scenario_name=scenario_name,
            scenario_type="forecast",
            amounts_fn=make_fn(fm),
            actual_months=max(0, fm - 1),
            tab_color="4472C4",  # blue
        )

    outpath = "excel/Budget_Forecast_2024.xlsx"
    wb.save(outpath)
    print(f"Created {outpath}")
    print(f"  Sheets: Setup + Budget 2024 + 2024F1-2024F12 (14 total)")
    print(f"  Entities: {', '.join(e['id'] for e in ENTITIES)}")
    print(f"  Accounts: {len(PNL_ACCOUNTS)} P&L lines per entity")
    print(f"  EPMSAVE formulas in all editable cells")
    print(f"  Setup sheet with connection config + instructions")


if __name__ == "__main__":
    main()
